# -*- coding: utf-8 -*-
"""
Author: 4wardEnergy Research GmbH
Date: 2024-05-14
Version: 1.0

Contains the gap-filling algorithms, including various modes 
of filling gaps in the consumer data. The script is designed to handle 
different methods for filling gaps such as using the last week's data, 
machine learning-based predictions, and standard load profiles.

Functions:
- print_red: Print text in red color.
- gaps: Main function to execute the gap-filling process based on selected mode.
- fill_mode_0: Default gap-filling mode which automatically selects an appropriate algorithm.
- fill_mode_1: Gap-filling by closing mass balance.
- fill_mode_2: Fill gaps using the last available value.
- fill_mode_3: Gap-filling using standard load profile.
- get_season: Determine the season for a given timestamp.
- close_vol_flow_dummy: Close volumetric flow balance for the gapfilling node.
- init_sim_data: Initialize simulation data for the current time step.
- fill_gap_last_week: Fill a gap with data from the previous week.
- fill_gap_ml: Fill a gap with machine learning-based predictions.
- fill_gap_slp: Fill a gap using standard load profiles.
"""
"""
Copyright (C) 2024  4wardEnergy Research GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import numpy as np
from openpyxl.styles import Font
from datetime import timedelta
import pandas as pd
import copy
import os
import calendar

import termcolor
from supervised.automl import AutoML

from simulation import fcns_gaps
from simulation import fcns_balance
from options import var_cons_prep, var_load_profiles, var_ml_models

def print_red(text):
    print(termcolor.colored(text, "red"))


def gaps (fileXLSX, fileXLSX_name, node, var_sim, balance, var_gaps):
    """Gapfilling algorithm for the data.

    :param fileXLSX: Active excel workbook.
    :type fileXLSX: openpyxl obj.

    :param fileXLSX_name: Path of the active workbook.
    :type fileXLSX_name: str

    :param node: Contains information about the nodes.
    :type node: node obj.

    :param var_sim: Variables concerning the simulation.
    :type var_sim: var_sim obj.

    :param balance: Balanced variables.
    :type balance: balance obj.

    :param var_gaps: Variables concerning the gapfilling.
    :type var_gaps: var_gaps obj.
    """    

    cntr = var_sim.cntr_time_hyd

###############################################################################
# GAPFILLING ##################################################################
###############################################################################

    if var_gaps.fill_mode == 0:
        fcns_gaps.fill_mode_0(node, var_sim, var_gaps, var_ml_models)

    elif var_gaps.fill_mode == 1:
        fcns_gaps.fill_mode_1(node, var_sim, balance, var_gaps)

    elif var_gaps.fill_mode == 2:
        fcns_gaps.fill_mode_2(node, var_sim)
    
    elif var_gaps.fill_mode == 3:
        fcns_gaps.fill_mode_3(node, var_sim, var_gaps)

    else:
        raise Exception(f"Gapfilling mode {var_gaps.fill_mode} not supported.")

    # Fill empty values with 0
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.cons[x_node] != None:
            if node.V_dot_sim[x_node][cntr] == None:
                node.V_dot_sim[x_node][cntr] = 0
            if node.Q_dot_sim[x_node][cntr] == None:
                node.Q_dot_sim[x_node][cntr] = 0

    # Check whether any node has the gapfilling node flag. If so, perform closing of volumetric flow balance for the gapfilling node.
    if cntr == 0:
        count_gapfilling_nodes = 0
        var_gaps.flag_close_vol_flow_dummy = False
        for x_node in range (0, node.nbr_orig_arabic.shape[0]):
                if node.gapfilling_node[x_node] == 1:
                    var_gaps.x_gapfilling_node = x_node
                    var_gaps.flag_close_vol_flow_dummy = True
                    count_gapfilling_nodes += 1
        if count_gapfilling_nodes > 1:
            raise Exception("More than one gapfilling node in the network. Please check the input file.")
    if var_gaps.flag_close_vol_flow_dummy:
        fcns_gaps.close_vol_flow_dummy(node, var_sim, var_gaps, cntr)
###############################################################################
# EXCEL #######################################################################
###############################################################################

    for x_node in range(0, node.nbr_orig_arabic.shape[0]):

        if node.cons[x_node] != None:

            # OPEN CORRESPONDING SHEET ########################################
            if node.cons[x_node] == "keine ID":
                sheet_cons = fileXLSX["C_"+str(node.nbr_orig_roman[x_node])]
            else:
                sheet_cons = fileXLSX["C_"+str(node.cons[x_node])]

            # ADD HEADER ######################################################
            if cntr == 0:
                sheet_cons_caption = ["V\u0307_sim [l/s]", "Q\u0307_sim [kW]"]
                for x in range (0, len(sheet_cons_caption)):
                    sheet_cons.cell(row = 1, column = x+10).value = \
                        sheet_cons_caption[x]

            # FILL EXCEL SHEET ############################################
            sheet_cons.cell(row = cntr+2, column = 10).value = node.V_dot_sim[x_node][cntr]
            sheet_cons.cell(row = cntr+2, column = 10).number_format = "0.000"
            sheet_cons.cell(row = cntr+2, column = 11).value = node.Q_dot_sim[x_node][cntr]
            sheet_cons.cell(row = cntr+2, column = 11).number_format = "0.0"

            # FORMATTING ######################################################
            bold_font = Font(bold = True)
            for cell in sheet_cons["1:1"]:
                cell.font = bold_font

    fileXLSX.save(fileXLSX_name)

# BOOKMARK: DEFAULT GAPFILLING MODE ###########################################
def fill_mode_0(node, var_sim, var_gaps, var_ml_models):
    """Gapfilling algorithm type 0: This is the standard mode. Gaps are filled
    using the method specified in the input excel file (Last week's data,
    ML-based or standard load profile).

    :param node: Contains information about the nodes.
    :type node: node obj.
    :param var_sim: Variables concerning the simulation.
    :type var_sim: var_sim obj.
    :param var_gaps: Variables concerning the gapfilling.
    :type var_gaps: var_gaps obj.
    :param var_ml_models: Variables concerning the ML models.
    :type var_ml_models: var_ml_models obj.
    """    

    cntr = var_sim.cntr_time_hyd

    # INITIALIZATION ##########################################################
    init_sim_data(node, cntr)

    # Check if SLP is used for any node
    flag_SLP = False
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.cons[x_node] != None:
            if node.gapfilling_mode[x_node] == "SLP":
                flag_SLP = True
    
    # If SLP is used, load weather data
    if flag_SLP:
        weather_data = pd.read_csv(var_load_profiles.weather_file)
        weather_data["time"] = weather_data["time"].str[:-6]
        weather_data["time"] = pd.to_datetime(weather_data["time"])
        weather_data["year"] = weather_data["time"].dt.year
        weather_data["month"] = weather_data["time"].dt.month
        weather_data["day"] = weather_data["time"].dt.day
        weather_data["hour"] = weather_data["time"].dt.hour

    # Loop through nodes
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.cons[x_node] != None:
            # Check for error flag
            if ("FEHLER" in node.error[x_node][cntr]) and (node.gapfilling_node[x_node] == 0):
                # Set gapfilling mode, override if necessary
                gapfilling_mode = node.gapfilling_mode[x_node]
                if node.gapfilling_override[x_node] is not None:
                    gapfilling_mode = node.gapfilling_override[x_node]

                # FILL GAP ACCORDING TO GAPFILLING MODE #######################
                if gapfilling_mode == "Vorwoche":
                    flag_couldnotfill = fill_gap_last_week(node, x_node, cntr, var_sim, var_gaps)
                elif gapfilling_mode == "ML":
                    flag_couldnotfill = fill_gap_ml(node, x_node, cntr, var_ml_models, var_sim, var_gaps)
                elif gapfilling_mode == "SLP":
                    flag_couldnotfill = fill_gap_slp(node, x_node, cntr, var_sim, var_gaps, weather_data)
                else:
                    raise Exception(f"Consumer {node.cons[x_node]}: Gap filling method {gapfilling_mode} not supported.")

                # CHECK IF GAP WAS FILLED ###################################
                if flag_couldnotfill:
                    print_red(f"Consumer {str(node.cons[x_node])}: Gap could not be filled with method {gapfilling_mode}. Filling with SLP.")
                    flag_couldnotfill = fill_gap_slp(node, x_node, cntr, var_sim, var_gaps, weather_data)
                if flag_couldnotfill:
                    print_red("Consumer " + str(node.cons[x_node]) + ": Gap could not be filled. Setting V_dot and Q_dot to 0.")
                    node.V_dot_sim[x_node][cntr] = 0
                    node.Q_dot_sim[x_node][cntr] = 0
            # Check if a number was returned
            if not isinstance(node.V_dot_sim[x_node][cntr], (int, float, np.float64)):
                raise Exception(f"Consumer {node.cons[x_node]}: V_dot_sim is not a number. Check input csv.")
            if not isinstance(node.Q_dot_sim[x_node][cntr], (int, float, np.float64)):
                raise Exception(f"Consumer {node.cons[x_node]}: Q_dot_sim is not a number. Check input csv.")

def fill_mode_1(node, var_sim, balance, var_gaps):
    """ Closing of mass balance.
        1. Go through all time steps of the simulation.
        2. If the volumetric flow balance is not closed at the current time step, 
           check for a consumer with an error flag.
        3. Add this consumers maximum power to Q_dot_max_error_sum.
        4. When all time steps are checked, go through all time steps again.
        5. For each consumer with an error Flag, set V_dot_corr to the delta of 
           the volumetric flow balance, weighted by the consumer's maximum power 
           in relation to Q_dot_error_sum. 

    :param node: Contains information about the nodes.
    :type node: node obj.
    :param var_sim: Variables concerning the simulation.
    :type var_sim: var_sim obj.
    :param balance: Balanced variables.
    :type balance: balance obj.
    :param var_gaps: Variables concerning the gapfilling.
    :type var_gaps: var_gaps obj.
    """    

    cntr = var_sim.cntr_time_hyd

    # INITIALIZATION ##########################################################
    init_sim_data(node, cntr)
    
    # Check if balance is not closed
    if balance.delta_V_dot[cntr] > 0:
        # CALCULATE SUM OF MAX. POWERS FOR CONSUMERS WITH ERROR FLAG ######
        Q_dot_max_error_sum = 0
        # Loop through nodes
        for x_node in range (0, node.nbr_orig_arabic.shape[0]):
            if node.cons[x_node] != None:
                # Check for error flag
                if "FEHLER" in node.error[x_node][cntr]:
                    # Add maximum power to sum
                    Q_dot_max_error_sum = Q_dot_max_error_sum+node.Q_dot_max[x_node]
        # DISTRIBUTE THE VOLUMETRIC FLOW GAP BETWEEN CONSUMERS ############
        for x_node in range (0, node.nbr_orig_arabic.shape[0]):
            if node.cons[x_node] != None:
                if node.gapfilling_node[x_node] == 1:
                    raise Exception("Gapfilling mode 1: Gapfilling dummy node cannot be used in this mode.")
                if "FEHLER" in node.error[x_node][cntr]:
                    if Q_dot_max_error_sum > 0:
                        node.V_dot_sim[x_node][cntr] = balance.delta_V_dot[cntr]*node.Q_dot_max[x_node]/Q_dot_max_error_sum
                        if (node.kWh_m3[x_node] != None) and (var_gaps.use_historical_kWh_m3):
                            node.Q_dot_sim[x_node][cntr] = node.V_dot_sim[x_node][cntr]*node.kWh_m3[x_node]*3600/1000
                        else:
                            node.Q_dot_sim[x_node][cntr] = node.V_dot_sim[x_node][cntr]*var_gaps.kWh_m3*3600/1000
                    else:
                        node.V_dot_sim[x_node][cntr] = 0
                        node.Q_dot_sim[x_node][cntr] = 0
    else:
        # FILL NANs WITH 0 EVEN WHEN THE BALANCE IS CLOSED ####################
        # This prevents errors from occuring at time steps where the losses
        # are negative.
        for x_node in range (0, node.nbr_orig_arabic.shape[0]):
            if node.cons[x_node] != None:
                if "FEHLER" in node.error[x_node][cntr]:
                    node.V_dot_sim[x_node][cntr] = 0
                    node.Q_dot_sim[x_node][cntr] = 0

def fill_mode_2(node, var_sim):
    """Fill gaps using the last available value.

    :param node: Contains information about the nodes.
    :type node: node obj.
    :param var_sim: Variables concerning the simulation.
    :type var_sim: var_sim obj.
    """

    cntr = var_sim.cntr_time_hyd

    # INITIALIZATION ##########################################################
    init_sim_data(node, cntr)

    # Loop through nodes
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.cons[x_node] != None:
            # Check for error flag
            if ("FEHLER" in node.error[x_node][cntr]) and (node.gapfilling_node[x_node] == 0):
                if cntr == 0:
                    raise Exception(f"Gapfilling mode 2: First time step is a gap for node {node.cons[x_node]}. Select another gapfilling mode or adjust simulation time frame.")
                # FILL GAP WITH LAST AVAILABLE VALUE ######################
                node.V_dot_sim[x_node][cntr] = node.V_dot_sim[x_node][cntr-1]
                node.Q_dot_sim[x_node][cntr] = node.Q_dot_sim[x_node][cntr-1]


def fill_mode_3(node, var_sim, var_gaps):
    """Gapfilling with standard load profile.

    :param node: Contains information about the nodes.
    :type node: node obj.
    :param var_sim: Variables concerning the simulation.
    :type var_sim: var_sim obj.
    :param var_gaps: Variables concerning the gapfilling.
    :type var_gaps: var_gaps obj.
    """

    cntr = var_sim.cntr_time_hyd

    ###########################################################################
    # INITIALIZATION ##########################################################
    ###########################################################################
    init_sim_data(node, cntr)

    weather_data = pd.read_csv(var_load_profiles.weather_file)
    weather_data["time"] = weather_data["time"].str[:-6]
    weather_data["time"] = pd.to_datetime(weather_data["time"])
    weather_data["year"] = weather_data["time"].dt.year
    weather_data["month"] = weather_data["time"].dt.month
    weather_data["day"] = weather_data["time"].dt.day
    weather_data["hour"] = weather_data["time"].dt.hour

    ###########################################################################
    # LOOP THROUGH ALL NODES ##################################################
    ###########################################################################
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        # CHECK IF THERE IS AN ERROR FLAG #####################################
        flag_error = False
        if node.cons[x_node] != None:
            if ("FEHLER" in node.error[x_node][cntr]) and (node.gapfilling_node[x_node] == 0):
                flag_error = True

        # IF THERE IS AN ERROR FLAG IN THIS NODE: PERFORM SLP GAPFILLING ######
        #######################################################################
        if flag_error:
            fill_gap_slp(node, x_node, cntr, var_sim, var_gaps, weather_data)


def get_season(time_stamp):
    """Returns the season of a given time stamp.

    :param time_stamp: time stamp to be analyzed.
    :type time_stamp: pd.Datetime
    :return: season of the given time stamp (0 = summer, 1 = winter, 2 = transition period)
    :rtype: int
    """
    year = time_stamp.year
    time_stamp = pd.to_datetime(time_stamp)
    start_summer = pd.Timestamp(year=year, month=5, day=15)
    end_summer = pd.Timestamp(year=year, month=9, day=14)
    start_winter = pd.Timestamp(year=year, month=11, day=1)
    end_winter = pd.Timestamp(year=year, month=3, day=20)
    if start_summer <= time_stamp <= end_summer:
        season = 0
    elif start_winter <= time_stamp <= end_winter:
        season = 1
    else:
        season = 2

    return season

def close_vol_flow_dummy(node, var_sim, var_gaps, cntr):
    """Closes the volumetric flow balance for the dummy node at 
    x_gapfilling_node.

    :param node: Contains information about the nodes.
    :type node: node obj.
    :param var_sim: Variables concerning the simulation.
    :type var_sim: var_sim obj.
    :param var_gaps: Variables concerning the gapfilling.
    :type var_gaps: var_gaps obj.
    """    
    
    x_gapfilling_node = var_gaps.x_gapfilling_node

    # Perform balancing
    delta_V_dot_sim = fcns_balance.balance_for_close_vol_flow_dummy(node, var_sim, cntr)

    # Check if balance is not closed. If that's the case, fill the gap.
    if delta_V_dot_sim > 0:
        node.V_dot_sim[x_gapfilling_node][cntr] = delta_V_dot_sim
        node.Q_dot_sim[x_gapfilling_node][cntr] = delta_V_dot_sim*var_gaps.kWh_m3*3600/1000
        node.error[x_gapfilling_node][cntr] = "Gapfilling-Knoten"


def init_sim_data(node, cntr):
    """Initializes the simulation data for the current time step.

    :param node: Contains information about the nodes.
    :type node: node obj.
    :param cntr: Current time step.
    :type cntr: int
    """    
    if cntr == 0:
        node.V_dot_sim = []
        node.Q_dot_sim = []
        for x_node in range (0, node.nbr_orig_arabic.shape[0]):
            if node.cons[x_node] != None:
                node.V_dot_sim.append([])
                node.Q_dot_sim.append([])
            else:
                node.V_dot_sim.append(None)
                node.Q_dot_sim.append(None)
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.cons[x_node] != None:
            node.V_dot_sim[x_node] = np.append(node.V_dot_sim[x_node],
                                               np.array([node.V_dot[x_node][cntr]]),
                                               axis = 0)
            node.Q_dot_sim[x_node] = np.append(node.Q_dot_sim[x_node],
                                               np.array([node.Q_dot[x_node][cntr]]),
                                               axis = 0)

# GAPFILLING ALGORITHMS FOR MODE 0 ############################################
###############################################################################

def fill_gap_last_week(node, x_node, cntr, var_sim, var_gaps):
    """ 
    Gapfilling algorithm for mode 0: Fill gap with last week's data.

    :param node: Contains information about the nodes.
    :type node: node obj.
    :param x_node: Index of the node.
    :type x_node: int
    :param cntr: Current time step.
    :type cntr: int
    :param var_sim: Variables concerning the simulation.
    :type var_sim: var_sim obj.
    :param var_gaps: Variables concerning the gapfilling.
    :type var_gaps: var_gaps obj.

    :return: flag_couldnotfill
    :rtype: bool
    """

    # Read csv file of consumer
    path = os.path.join(var_cons_prep.cons_dir, "Regler_" + str(node.cons[x_node]) + "_prepared.csv")
    df = pd.read_csv(path)
    # Convert "Zeitstempel" to datetime
    df["Zeitstempel"] = pd.to_datetime(df["Zeitstempel"])

    # Find current timestep in dataframe
    current_time = pd.to_datetime(var_sim.time_sim)
    # Go 7 days back
    last_week = current_time - timedelta(days=7)

    # Filter dataframe for timestamp of last week
    try:
        df_filt = df.loc[df["Zeitstempel"] <= last_week]
    except:
        flag_couldnotfill = True

    # Check if data is invalid (forerun temp. has not been changing for more than var_gaps.nbr_equal_values_max time steps)
    # Filter the last var_gaps.nbr_equal_values_max time steps
    df_filt = df_filt.tail(int(np.round(var_gaps.equal_values_max_min/var_sim.delta_time_hyd)))
    # Check if the temperature has been changing
    if df_filt["Vorlauftemperatur (°C)"].nunique() == 1:
        print("Datenausfall in der Vorwoche.")
        flag_couldnotfill = True
    else:
        flag_couldnotfill = False
    
    # Filter for the last time step
    df_filt = df_filt.tail(1)
    df_filt = df_filt.reset_index()

    # Check if the filtered dataframe is empty
    if df_filt.empty:
        flag_couldnotfill = True
    
    # Check if the date of df_filt is the same as last_week
    if df_filt["Zeitstempel"][0] != last_week:
        flag_couldnotfill = True

    # Read "akt. Leistung(kW)" and "Durchfluss (l/h)" from dataframe
    V_dot = df_filt["Durchfluss (l/h)"].values[0]/3600
    Q_dot = df_filt["akt. Leistung(kW)"].values[0] 

    # Check if the values are floats or ints
    if (type(V_dot) != int) and (type(V_dot) != float) and (type(V_dot) != np.float64):
        flag_couldnotfill = True
    if (type(Q_dot) != int) and (type(Q_dot) != float) and (type(V_dot) != np.float64):
        flag_couldnotfill = True

    # Fill the gap
    node.V_dot_sim[x_node][cntr] = V_dot
    node.Q_dot_sim[x_node][cntr] = Q_dot  

    return flag_couldnotfill

def fill_gap_ml(node, x_node, cntr, var_ml_models, var_sim, var_gaps):
    """
    Gapfilling algorithm for mode 0: Fill gap with ML-based prediction.

    :param node: Contains information about the nodes.
    :type node: node obj.
    :param x_node: Index of the node.
    :type x_node: int
    :param cntr: Current time step.
    :type cntr: int
    :param var_ml_models: Variables concerning the ML models.
    :type var_ml_models: var_ml_models obj.
    :param var_sim: Variables concerning the simulation.
    :type var_sim: var_sim obj.
    :param var_gaps: Variables concerning the gapfilling.
    :type var_gaps: var_gaps obj.

    :returns: flag_couldnotfill
    :rtype: bool
    """
    flag_couldnotfill = False

    # Set model path
    cons_id = node.cons[x_node]
    model_path = os.path.join(var_ml_models.ml_model_dir, cons_id, "AutoML_1")

    # Load model
    try:
        automl = AutoML(results_path=model_path)
    except:
        print_red(f"ML model for consumer {cons_id} not found.")
        flag_couldnotfill = True

    # Predict missing timestamp
    current_time = pd.to_datetime(var_sim.time_sim)
    try:
        Q_dot = automl.predict(current_time)
    except:
        print_red(f"ML gapfilling failed for consumer {cons_id}.")
        flag_couldnotfill = True

    # Check if there's a historical value available for kWh_m3
    if (node.kWh_m3[x_node] != None) and (var_gaps.use_historical_kWh_m3):
        V_dot = Q_dot*1000/(node.kWh_m3[x_node]*3600)
    # If not, use standard value
    else:
        V_dot = Q_dot*1000/(var_gaps.kWh_m3*3600)

    #Check if Q_dot is < 0; if so, set to 0
    if Q_dot < 0:
        Q_dot = 0
        V_dot = 0

    # Fill the gap
    node.V_dot_sim[x_node][cntr] = V_dot
    node.Q_dot_sim[x_node][cntr] = Q_dot 

    return flag_couldnotfill


def fill_gap_slp(node, x_node, cntr, var_sim, var_gaps, weather_data):
    """
    Gapfilling algorithm for modes 0 and 3: Fill gap with standard load profile.
    Fills the gap for a given node and time step with heating and SHW load profiles.

    :param node: The node object.
    :type node: Node
    :param x_node: The index of the node.
    :type x_node: int
    :param cntr: The index of the time step.
    :type cntr: int
    :param var_sim: The simulation variables object.
    :type var_sim: VarSim
    :param var_gaps: The gap filling variables object.
    :type var_gaps: VarGaps
    :param weather_data: The weather data.
    :type weather_data: pd.DataFrame

    :returns: None
    :rtype: None
    """

    time_loop = var_sim.time_sim

    # Check if the consumer is residential
    flag_shw = "no"
    if "wohn" in node.building_type[x_node]:
        # Check if there is historical data available
        if node.hist_data_available[x_node]:
            # If there is: Read historical data, mark for SHW filling
            path_hist_data = os.path.join(var_cons_prep.cons_dir, "Regler_" + str(node.cons[x_node]) + "_prepared.csv")
            shw_hist_data = pd.read_csv(path_hist_data)

            # Extract time information
            shw_hist_data["Zeitstempel"] = pd.to_datetime(shw_hist_data["Zeitstempel"])
            shw_hist_data["Jahr"] = shw_hist_data["Zeitstempel"].dt.year
            shw_hist_data["Kalenderwoche"] = shw_hist_data["Zeitstempel"].dt.isocalendar().week
            shw_hist_data["Wochentag"] = shw_hist_data["Zeitstempel"].dt.weekday
            shw_hist_data["Stunde"] = shw_hist_data["Zeitstempel"].dt.hour
            shw_hist_data["Minute"] = shw_hist_data["Zeitstempel"].dt.minute

            flag_shw = "yes, hist"
        else:
            # If there is no hist. data: Load load profile, mark for SHW filling
            path_slp = os.path.join(var_load_profiles.load_profile_dir, "individual", "Regler_"+ str(node.cons[x_node]) + "_shw.csv")
            shw_slp = pd.read_csv(path_slp)
            flag_shw = "yes, lp"
    
    # Load heating load profile
    path_slp = os.path.join(var_load_profiles.load_profile_dir, "individual", "Regler_"+ str(node.cons[x_node]) + "_heating.csv")
    heating_slp = pd.read_csv(path_slp)
    
    # Check if "FEHLER" is in node.error
    if "FEHLER" in node.error[x_node][cntr]:
        power_heating = 0
        power_slp = 0
        
        # SHW FILLING
        if "yes" in flag_shw:
            if "hist" in flag_shw:
                # GAPFILLING WITH HISTORICAL DATA
                # Extract info from the current time step
                year = time_loop.year - 1
                week = time_loop.isocalendar()[1]
                weekday = time_loop.weekday()
                hour = time_loop.hour
                minute = time_loop.minute
                # Select matching data from historical data
                matching_entry = shw_hist_data.loc[(shw_hist_data["Jahr"] == year) & (shw_hist_data["Kalenderwoche"] == week) & \
                                                (shw_hist_data["Wochentag"] == weekday) & (shw_hist_data["Stunde"] == hour) & \
                                                (shw_hist_data["Minute"] == minute)]
                power_shw = matching_entry["Warmwasser (kW)"].values[0]
            else:
                # GAPFILLING WITH LOAD PROFILE
                # Extract season info from the current time step
                season = get_season(time_loop)
                # Extract type of day from the current time step
                daytype = time_loop.weekday()
                daytype = 0 if daytype <= 4 else 1 if daytype == 5 else 2
                # Extract hour from the current time step
                hour = time_loop.hour
                # Load profile hour naming conventions are different from the ones used in pandas
                lp_hour = 24 if hour == 0 else hour
                # Select matching data from load profile
                matching_entry = shw_slp.loc[(shw_slp["season"] == season) & (shw_slp["daytype"] == daytype) & \
                                                (shw_slp["hour"] == lp_hour)]
                power_shw = matching_entry["load"].values[0]
        else:
            power_shw = 0
        
        # HEATING FILLING
        # EXTRACT INFO FROM TIMESTAMP
        daytype = time_loop.weekday()
        daytype = 0 if daytype <= 4 else 1 if daytype == 5 else 2  

        year = time_loop.year
        month = time_loop.month
        day = time_loop.day
        hour = time_loop.hour
        # Load profile hour naming conventions are different from the ones used in pandas
        lp_hour = 24 if hour == 0 else hour

        outside_temp = weather_data.loc[(weather_data["year"] == year) & (weather_data["month"] == month) & \
                                        (weather_data["day"] == day) & (weather_data["hour"] == hour)]["TTX"].values[0]
        outside_temp = round(outside_temp)
        # Set to minimum value incl. in load profile in case it subcedes it
        outside_temp = -15 if outside_temp < -15 else outside_temp

        # FILL WITH SLP
        # INDUSTRIAL HEATING
        if "ind" in node.building_type[x_node]:
            # Match type of day, month and hour
            matching_entry = heating_slp.loc[(heating_slp["daytype"] == daytype) & (heating_slp["month"] == month) & \
                                                (heating_slp["hour"] == lp_hour)]
            power_heating = matching_entry["load"].values[0]
        # TERTIARY HEATING
        elif ("tert" in node.building_type[x_node]) or ("wohn" in node.building_type[x_node]):
            if outside_temp <= 17:
                # Match type of day, outside temp. and hour
                matching_entry = heating_slp.loc[(heating_slp["daytype"] == daytype) & (heating_slp["temperature"] == outside_temp) & \
                                                (heating_slp["hour"] == lp_hour)]
                power_heating = matching_entry["load"].values[0]
            else:
                power_heating = 0
        # CATCH OTHER
        else:
            raise Exception(f"Gapfilling mode 3: Building type {node.building_type[x_node]} not supported.")

        # SUM UP POWER AND WRITE INTO Q_dot_sim
        power_sum = power_shw + power_heating
        node.Q_dot_sim[x_node][cntr] = power_sum
        # Calculate V_dot_sim from Q_dot_sim
        # Check if there's a historical value available for kWh_m3
        if (node.kWh_m3[x_node] != None) and (var_gaps.use_historical_kWh_m3):
            node.V_dot_sim[x_node][cntr] = power_sum*1000/(node.kWh_m3[x_node]*3600)
        # If not, use standard value
        else:
            node.V_dot_sim[x_node][cntr] = power_sum*1000/(var_gaps.kWh_m3*3600)

    return False
