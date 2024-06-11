# -*- coding: utf-8 -*-
"""
Author: 4wardEnergy Research GmbH
Date: 2024-05-14
Version: 1.0

This script contains functions for performing and recording the balancing of measurements 
and simulated values within the network. The balances are recorded to the output Excel file.

Functions:
- balance_meas: Balances the measurements and records the results in an Excel file.
- balance_sim: Balances the simulation results after gap filling and records the results in an Excel file.
- balance_for_close_vol_flow_dummy: Balances volume flow for the gapfilling node.
"""
"""
Copyright (C) 2024  4wardEnergy Research GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from openpyxl.styles import Font
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import NamedStyle
from statistics import mean
from simulation import fcns_phy

def balance_meas(var_misc, var_sim, fileXLSX, fileXLSX_name, node, balance):
    """Performs balancing of the measurements.

    :param var_misc: Miscellaneous variables
    :type var_misc: var_misc obj.

    :param var_sim: Variables concerning the simulation
    :type var_sim: var_sim obj.

    :param fileXLSX: The active workbook
    :type fileXLSX: openpyxl.workbook.workbook obj.

    :param fileXLSX_name: Path of the active workbook
    :type fileXLSX_name: str

    :param node: Contains information about the nodes
    :type node: node obj.

    :return: Balanced variables
    :rtype: balance obj.
    """    

    cntr = var_sim.cntr_time_hyd

###############################################################################
# BALANCING ###################################################################
###############################################################################

    # INITIALIZATION #########################################################
    delta_V_dot, delta_V_dot_pct, delta_m_dot, delta_m_dot_pct, delta_Q_dot, \
        delta_Q_dot_pct, Q_dot_error_sum, Q_year_error_sum, \
        H2O_year_error_sum = ([] for i in range(9))

    V_dot_sum_int, V_dot_feed_sum_int, m_dot_sum_int, m_dot_feed_sum_int, \
        Q_dot_sum_int, Q_dot_feed_sum_int, Q_dot_error_sum_int, \
        Q_year_error_sum_int, H2O_year_error_sum_int \
        = (0 for i in range(9))
    
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):

        # SUMMATION OF CONSUMERS ##########################################
        if not isinstance(node.V_dot[x_node], type(None)): 
            if node.error[x_node][cntr] == "OKAY_m" or \
                node.error[x_node][cntr] == "OKAY_a":

                # Summation of measurements
                V_dot_sum_int = V_dot_sum_int+node.V_dot[x_node][cntr]
                m_dot_sum_int = m_dot_sum_int+\
                    node.V_dot[x_node][cntr]*10**(-3)*\
                    fcns_phy.H2O_density(mean([node.\
                    temp_flow[x_node][cntr], node.\
                    temp_ret[x_node][cntr]]))
                Q_dot_sum_int = Q_dot_sum_int+\
                    node.Q_dot[x_node][cntr]
            else:

                # summation of stored data of the transfer stations
                # in case of measurement failure
                Q_dot_error_sum_int = Q_dot_error_sum_int+\
                    (node.Q_dot_max[x_node] if isinstance(node.Q_dot_max[x_node], (int,float)) else 0)
                Q_year_error_sum_int = Q_year_error_sum_int+\
                    (node.Q_year[x_node] if isinstance(node.Q_year[x_node], (int,float)) else 0)
                H2O_year_error_sum_int = H2O_year_error_sum_int+\
                    (node.H2O_year[x_node] if isinstance(node.H2O_year[x_node], (int,float)) else 0)

        # SUMMATION OF FEEDERS ############################################
        if not isinstance(node.V_dot_feed[x_node], type(None)):
            if node.error_feed[x_node][cntr] == "OKAY_m" or \
                node.error_feed[x_node][cntr] == "OKAY_a":
                V_dot_feed_sum_int = V_dot_feed_sum_int+\
                    node.V_dot_feed[x_node][cntr]
                m_dot_feed_sum_int = m_dot_feed_sum_int+\
                    node.V_dot_feed[x_node][cntr]*10**(-3)*\
                    fcns_phy.H2O_density(mean([node.\
                    temp_flow_feed[x_node][cntr], node.\
                    temp_ret_feed[x_node][cntr]]))
                Q_dot_feed_sum_int = Q_dot_feed_sum_int+\
                    node.Q_dot_feed[x_node][cntr]

    # BALANCING ###########################################################
    # for volume-, mass- and heat flow
    delta_V_dot.append(V_dot_feed_sum_int-V_dot_sum_int)
    delta_m_dot.append(m_dot_feed_sum_int-m_dot_sum_int)
    delta_Q_dot.append(Q_dot_feed_sum_int-Q_dot_sum_int)

    # Calculate percentages
    if V_dot_feed_sum_int != 0:
        delta_V_dot_pct.append(delta_V_dot[-1]/V_dot_feed_sum_int*100)
    else:
        delta_V_dot_pct.append("-")
    if m_dot_feed_sum_int != 0:
        delta_m_dot_pct.append(delta_m_dot[-1]/m_dot_feed_sum_int*100)
    else:
        delta_m_dot_pct.append("-")
    if Q_dot_feed_sum_int != 0:
        delta_Q_dot_pct.append(delta_Q_dot[-1]/Q_dot_feed_sum_int*100)
    else:
        delta_Q_dot_pct.append("-")

    Q_dot_error_sum.append(Q_dot_error_sum_int)
    Q_year_error_sum.append(Q_year_error_sum_int)
    H2O_year_error_sum.append(H2O_year_error_sum_int)

###############################################################################
# EXCEL (VOLUME) ##############################################################
###############################################################################

    # CREATE SHEET ############################################################
    if cntr == 0:
        fileXLSX.create_sheet("V_dot_Bilanz_m")
        fileXLSX.save(fileXLSX_name)
    sheet_delta_V_dot_m = fileXLSX["V_dot_Bilanz_m"]

    # TIME STAMPS #############################################################
    if cntr == 0:
        sheet_delta_V_dot_m.cell(row = 1, column = 1).value = \
            "Zeit ↓ Nr. → V\u0307 [l/s] ↘"

    sheet_delta_V_dot_m.cell(row = cntr+2, column = 1).value = \
        var_sim.time_stamp[cntr]
    sheet_delta_V_dot_m.cell(row = cntr+2, column = 1).style = \
        var_misc.date_style

    # CONSUMERS ###############################################################
    x_column = 2
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.cons[x_node] != None:

            # CONSUMERS WITHOUT ID ############################################
            if node.cons[x_node] == "keine ID":

                # Überschrift
                if cntr == 0:
                    sheet_delta_V_dot_m.cell(row = 1, column = x_column).value = \
                        node.nbr_orig_roman[x_node]

                # Werte

                sheet_delta_V_dot_m.cell(row = cntr+2, column = \
                                            x_column).value = "-"

            # CONSUMERS WITH ID ###############################################
            else:

                # Überschrift
                if cntr == 0:
                    sheet_delta_V_dot_m.cell(row = 1, column = x_column).value = \
                        node.cons[x_node]

                # Werte
                if node.error[x_node][cntr] == "OKAY_m" or \
                    node.error[x_node][cntr] == "OKAY_a":
                    sheet_delta_V_dot_m.cell(row = cntr+2, column = \
                        x_column).value = node.V_dot[x_node][cntr]
                    sheet_delta_V_dot_m.cell(row = cntr+2, column = \
                        x_column).number_format = "0.000"

            x_column += 1

    # FEEDERS #################################################################
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.feed_in[x_node] != None:

            # Header
            if cntr == 0:
                sheet_delta_V_dot_m.cell(row = 1, column = x_column).value = \
                    "F_"+str(node.feed_in[x_node])

            # Values
            if node.error_feed[x_node][cntr] == "OKAY_m" or \
                node.error_feed[x_node][cntr] == "OKAY_a":
                sheet_delta_V_dot_m.cell(row = cntr+2, column = \
                    x_column).value = node.V_dot_feed[x_node][cntr]
                sheet_delta_V_dot_m.cell(row = cntr+2, column = \
                    x_column).number_format = "0.000"

            x_column += 1

    # BALANCING ###############################################################
    # Header
    if cntr == 0:
        sheet_delta_V_dot_m.cell(row = 1, column = x_column).value = \
            "\u0394V\u0307_m [l/s]"
        sheet_delta_V_dot_m.cell(row = 1, column = x_column+1).value = \
            "\u0394V\u0307_m [%]"
        sheet_delta_V_dot_m.cell(row = 1, column = x_column+2).value = \
            "\u0394m\u0307_m [kg/s]"
        sheet_delta_V_dot_m.cell(row = 1, column = x_column+3).value = \
            "\u0394m\u0307_m [%]"

    # Values
    sheet_delta_V_dot_m.cell(row = cntr+2, column = x_column).value = \
        delta_V_dot[0]
    sheet_delta_V_dot_m.cell(row = cntr+2, column = \
        x_column).number_format = "0.000"
    sheet_delta_V_dot_m.cell(row = cntr+2, column = x_column+1).value = \
        delta_V_dot_pct[0]
    sheet_delta_V_dot_m.cell(row = cntr+2, column = \
        x_column+1).number_format = "0.00"
    sheet_delta_V_dot_m.cell(row = cntr+2, column = x_column+2).value = \
        delta_m_dot[0]
    sheet_delta_V_dot_m.cell(row = cntr+2, column = \
        x_column+2).number_format = "0.000"
    sheet_delta_V_dot_m.cell(row = cntr+2, column = x_column+3).value = \
        delta_m_dot_pct[0]
    sheet_delta_V_dot_m.cell(row = cntr+2, column = \
        x_column+3).number_format = "0.00"

    # FORMATTING ##############################################################
    bold_font = Font(bold = True)
    for cell in sheet_delta_V_dot_m["1:1"]:
        cell.font = bold_font
    sheet_delta_V_dot_m.freeze_panes = 'A2'
    sheet_delta_V_dot_m.column_dimensions["A"].width = 20

    # DIAGRAM #################################################################
    # Volume
    # Delete old charts
    if cntr != 0:
        for chart in sheet_delta_V_dot_m._charts:
            del chart

    chart = ScatterChart() # Excel-Diagramm-Klasse
    chart.title = "Volumenstromverluste über das Netz"
    chart.style = 13
    chart.x_axis.title = "Zeit"
    chart.x_axis.number_format = "hh:mm:ss"
    chart.y_axis.title = "Volumenstromverlust [%]"
    chart.width = 30

    x_values = Reference(sheet_delta_V_dot_m, min_col = 1, min_row = 2, \
                        max_row = len(var_sim.time_stamp)+1)
    y_values = Reference(sheet_delta_V_dot_m, min_col = x_column+1, \
                        min_row = 2, max_row = len(var_sim.time_stamp)+1)

    series = Series(y_values, x_values, title_from_data = False)
    chart.series.append(series)
    chart.legend = None

    sheet_delta_V_dot_m.add_chart(chart, "A13")

    # Mass
    chart = ScatterChart()
    chart.title = "Massenstromverluste über das Netz"
    chart.style = 13
    chart.x_axis.title = "Zeit"
    chart.x_axis.number_format = "hh:mm:ss"
    chart.y_axis.title = "Massenstromverlust [%]"
    chart.width = 30

    x_values = Reference(sheet_delta_V_dot_m, min_col = 1, min_row = 2, \
                        max_row = len(var_sim.time_stamp)+1)
    y_values = Reference(sheet_delta_V_dot_m, min_col = x_column+3, \
                        min_row = 2, max_row = len(var_sim.time_stamp)+1)

    series = Series(y_values, x_values, title_from_data = False)
    chart.series.append(series)
    chart.legend = None

    sheet_delta_V_dot_m.add_chart(chart, "A28")

    fileXLSX.save(fileXLSX_name)

###############################################################################
# EXCEL (POWER) ###############################################################
###############################################################################

    # CREATE SHEET ############################################################
    if cntr == 0:
        fileXLSX.create_sheet("Q_dot_Bilanz_m")
        fileXLSX.save(fileXLSX_name)
    sheet_delta_Q_dot_m = fileXLSX["Q_dot_Bilanz_m"]

    # TIME STAMPS #############################################################
    if cntr == 0:
        sheet_delta_Q_dot_m.cell(row = 1, column = 1).value = \
            "Zeit ↓ Nr. → Q\u0307 [kW] ↘"
        
    sheet_delta_Q_dot_m.cell(row = cntr+2, column = 1).value = \
        var_sim.time_stamp[cntr]
    sheet_delta_Q_dot_m.cell(row = cntr+2, column = 1).style = \
        var_misc.date_style

    # CONSUMERS ###############################################################
    x_column = 2
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.cons[x_node] != None:

            # CONSUMERS WITHOUT ID ############################################
            if node.cons[x_node] == "keine ID":

                # Header
                if cntr == 0:
                    sheet_delta_Q_dot_m.cell(row = 1, column = x_column).value = \
                        node.nbr_orig_roman[x_node]

                # Values
                sheet_delta_Q_dot_m.cell(row = cntr+2, column = \
                                            x_column).value = "-"

            # CONSUMERS WITH ID #################################################
            else:

                # Header
                if cntr == 0:
                    sheet_delta_Q_dot_m.cell(row = 1, column = x_column).value = \
                        node.cons[x_node]

                # Values
                if node.error[x_node][cntr] == "OKAY_m" or \
                    node.error[x_node][cntr] == "OKAY_a":
                    sheet_delta_Q_dot_m.cell(row = cntr+2, column = \
                        x_column).value = node.Q_dot[x_node][cntr]

            x_column += 1

    # FEEDERS #################################################################

    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.feed_in[x_node] != None:

            # Header
            if cntr == 0:
                sheet_delta_Q_dot_m.cell(row = 1, column = x_column).value = \
                    "F_"+str(node.feed_in[x_node])

            # Values
            if node.error_feed[x_node][cntr] == "OKAY_m" or \
                node.error_feed[x_node][cntr] == "OKAY_a":
                sheet_delta_Q_dot_m.cell(row = cntr+2, column = \
                    x_column).value = node.Q_dot_feed[x_node][cntr]

            x_column += 1

    # BALANCING ###############################################################
    # Header
    if cntr == 0:
        sheet_delta_Q_dot_m.cell(row = 1, column = x_column).value = \
            "\u0394Q\u0307_m [kW]"
        sheet_delta_Q_dot_m.cell(row = 1, column = x_column+1).value = \
            "\u0394Q\u0307_m [%]"

    # Values
    sheet_delta_Q_dot_m.cell(row = cntr+2, column = x_column).value = \
        delta_Q_dot[0]
    sheet_delta_Q_dot_m.cell(row = cntr+2, column = x_column+1).value = \
        delta_Q_dot_pct[0]
    sheet_delta_Q_dot_m.cell(row = cntr+2, column = \
        x_column+1).number_format = "0.00"

    # FORMATTING ##############################################################
    bold_font = Font(bold = True)
    for cell in sheet_delta_Q_dot_m["1:1"]:
        cell.font = bold_font
    sheet_delta_Q_dot_m.freeze_panes = 'A2'
    sheet_delta_Q_dot_m.column_dimensions["A"].width = 20

    # DIAGRAM #################################################################
    # Delete old charts
    if cntr != 0:
        for chart in sheet_delta_Q_dot_m._charts:
            del chart
            
    chart = ScatterChart()
    chart.title = "Wärmeverluste über das Netz"
    chart.style = 13
    chart.x_axis.title = "Zeit"
    chart.x_axis.number_format = "hh:mm:ss"
    chart.y_axis.title = "Wärmeverlust [%]"
    chart.width = 30

    x_values = Reference(sheet_delta_Q_dot_m, min_col = 1, min_row = 2, \
                        max_row = len(var_sim.time_stamp)+1)
    y_values = Reference(sheet_delta_Q_dot_m, min_col = x_column+1, \
                        min_row = 2, max_row = len(var_sim.time_stamp)+1)

    series = Series(y_values, x_values, title_from_data = False)
    chart.series.append(series)
    chart.legend = None

    sheet_delta_Q_dot_m.add_chart(chart, "A13")

    fileXLSX.save(fileXLSX_name)

###############################################################################
# OUTPUT ######################################################################
###############################################################################
    if cntr == 0:
        balance.delta_V_dot = []
        balance.delta_m_dot = []
        balance.delta_Q_dot = []
        balance.delta_V_dot_pct = []
        balance.delta_Q_dot_pct = []
        balance.Q_dot_error_sum = []
        balance.Q_year_error_sum = []
        balance.H2O_year_error_sum = []

    balance.delta_V_dot.append(delta_V_dot[0])
    balance.delta_m_dot.append(delta_m_dot[0])
    balance.delta_Q_dot.append(delta_Q_dot[0])
    balance.delta_V_dot_pct.append(delta_V_dot[0])
    balance.delta_Q_dot_pct.append(delta_Q_dot[0])
    balance.Q_dot_error_sum.append(Q_dot_error_sum[0])
    balance.Q_year_error_sum.append(Q_year_error_sum[0])
    balance.H2O_year_error_sum.append(H2O_year_error_sum[0])

    return (balance)

def balance_sim(var_misc, var_sim, fileXLSX, fileXLSX_name, node, balance):
    """Performs balancing after gapfilling.

    :param var_misc: Miscellaneous variables
    :type var_misc: var_misc obj.

    :param var_sim: Variables concerning the simulation
    :type var_sim: var_sim obj.

    :param fileXLSX: The active workbook
    :type fileXLSX: openpyxl.workbook.workbook obj.

    :param fileXLSX_name: Path of the active workbook
    :type fileXLSX_name: str

    :param node: Contains information about the nodes
    :type node: node obj.
    """  

    cntr = var_sim.cntr_time_hyd

###############################################################################
# BALANCING ###################################################################
###############################################################################

    # INITIALIZATION ##########################################################
    delta_V_dot_sim, delta_V_dot_sim_pct, delta_m_dot_sim, \
        delta_m_dot_sim_pct = ([] for i in range(4))

    V_dot_sum_int, V_dot_feed_sum_int, m_dot_sum_int, m_dot_feed_sum_int \
        = (0 for i in range(4))

    for x_node in range (0, node.nbr_orig_arabic.shape[0]):

        # SUMMATION OF CONSUMERS ##########################################
        if not isinstance(node.V_dot_sim[x_node], type(None)):

            # Summation of measurements
            V_dot_sum_int = V_dot_sum_int+node.V_dot_sim[x_node][cntr]
            if node.temp_flow[x_node][cntr] == None or \
                node.temp_ret[x_node][cntr] == None:
                m_dot_sum_int = m_dot_sum_int+\
                    node.V_dot_sim[x_node][cntr]*10**(-3)*\
                    fcns_phy.H2O_density(50)
            else:
                m_dot_sum_int = m_dot_sum_int+\
                    node.V_dot_sim[x_node][cntr]*10**(-3)*\
                    fcns_phy.H2O_density(mean([node.\
                    temp_flow[x_node][cntr], node.\
                    temp_ret[x_node][cntr]]))

        # SUMMATION OF FEEDERS ############################################
        if not isinstance(node.V_dot_feed[x_node], type(None)):
            V_dot_feed_sum_int = V_dot_feed_sum_int+\
                node.V_dot_feed[x_node][cntr]
            m_dot_feed_sum_int = m_dot_feed_sum_int+\
                node.V_dot_feed[x_node][cntr]*10**(-3)*\
                fcns_phy.H2O_density(mean([node.\
                temp_flow_feed[x_node][cntr], node.\
                temp_ret_feed[x_node][cntr]]))

    # BALANCING ###########################################################
    delta_V_dot_sim.append(V_dot_feed_sum_int-V_dot_sum_int)
    delta_m_dot_sim.append(m_dot_feed_sum_int-m_dot_sum_int)

    if V_dot_feed_sum_int != 0:
        delta_V_dot_sim_pct.append(delta_V_dot_sim[-1]/V_dot_feed_sum_int*100)
    else:
        delta_V_dot_sim_pct.append("-")
    if m_dot_feed_sum_int != 0:
        delta_m_dot_sim_pct.append(delta_m_dot_sim[-1]/m_dot_feed_sum_int*100)
    else:
        delta_m_dot_sim_pct.append("-")

###############################################################################
# EXCEL (VOLUME) ##############################################################
###############################################################################

    # CREATE SHEET ############################################################
    if cntr == 0:
        fileXLSX.create_sheet("V_dot_Bilanz_sim")
        fileXLSX.save(fileXLSX_name)
    sheet_delta_V_dot_sim = fileXLSX["V_dot_Bilanz_sim"]

    # TIME STAMPS #############################################################
    if cntr == 0:
        sheet_delta_V_dot_sim.cell(row = 1, column = 1).value = \
            "Zeit ↓ Nr. → V\u0307 [l/s] ↘"

    sheet_delta_V_dot_sim.cell(row = cntr+2, column = 1).value = \
        var_sim.time_stamp[cntr]
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = 1).style = \
        var_misc.date_style

    # CONSUMERS ###############################################################
    x_column = 2
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.cons[x_node] != None:

            # CONSUMERS WITHOUT ID ############################################
            if node.cons[x_node] == "keine ID":

                # Header
                if cntr == 0:
                    sheet_delta_V_dot_sim.cell(row = 1, column = x_column).value = \
                        node.nbr_orig_roman[x_node]

                # Values
                sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
                    x_column).value = node.V_dot_sim[x_node][cntr]
                sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
                    x_column).number_format = "0.000"

            # CONSUMERS WITH ID ###############################################
            else:

                # Überschrift
                if cntr == 0:
                    sheet_delta_V_dot_sim.cell(row = 1, column = x_column).value = \
                        node.cons[x_node]

                # Werte
                sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
                    x_column).value = node.V_dot_sim[x_node][cntr]
                sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
                    x_column).number_format = "0.000"

            x_column += 1

    # FEEDERS #################################################################
    for x_node in range (0, node.nbr_orig_arabic.shape[0]):
        if node.feed_in[x_node] != None:

            # Header
            if cntr == 0:
                sheet_delta_V_dot_sim.cell(row = 1, column = x_column).value = \
                    "F_"+str(node.feed_in[x_node])

            # Values
            if node.error_feed[x_node][cntr] == "OKAY_m" or \
                node.error_feed[x_node][cntr] == "OKAY_a":
                sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
                    x_column).value = node.V_dot_feed[x_node][cntr]
                sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
                    x_column).number_format = "0.000"

            x_column += 1

    # BALANCING ###############################################################
    # Header
    if cntr == 0:
        sheet_delta_V_dot_sim.cell(row = 1, column = x_column).value = \
            "\u0394V\u0307_sim [l/s]"
        sheet_delta_V_dot_sim.cell(row = 1, column = x_column+1).value = \
            "\u0394V\u0307_sim [%]"
        sheet_delta_V_dot_sim.cell(row = 1, column = x_column+2).value = \
            "\u0394m\u0307_sim [kg/s]"
        sheet_delta_V_dot_sim.cell(row = 1, column = x_column+3).value = \
            "\u0394m\u0307_sim [%]"

    # Values
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = x_column).value = \
        delta_V_dot_sim[0]
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
        x_column).number_format = "0.000"
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = x_column+1).value = \
        delta_V_dot_sim_pct[0]
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
        x_column+1).number_format = "0.00"
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = x_column+2).value = \
        delta_m_dot_sim[0]
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
        x_column+2).number_format = "0.000"
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = x_column+3).value = \
        delta_m_dot_sim_pct[0]
    sheet_delta_V_dot_sim.cell(row = cntr+2, column = \
        x_column+3).number_format = "0.00"

    # FORMATTING ##############################################################
    bold_font = Font(bold = True)
    for cell in sheet_delta_V_dot_sim["1:1"]:
        cell.font = bold_font
    sheet_delta_V_dot_sim.freeze_panes = 'A2'
    sheet_delta_V_dot_sim.column_dimensions["A"].width = 20

    # DIAGRAM #################################################################
    # Volume
    # Delete old charts
    if cntr != 0:
        for chart in sheet_delta_V_dot_sim._charts:
            del chart
    chart = ScatterChart()
    chart.title = "Volumenstromverluste über das Netz"
    chart.style = 13
    chart.x_axis.title = "Zeit"
    chart.x_axis.number_format = "hh:mm:ss"
    chart.y_axis.title = "Volumenstromverlust [%]"
    chart.width = 30

    x_values = Reference(sheet_delta_V_dot_sim, min_col = 1, min_row = 2, \
                        max_row = len(var_sim.time_stamp)+1)
    y_values = Reference(sheet_delta_V_dot_sim, min_col = x_column+1, \
                        min_row = 2, max_row = len(var_sim.time_stamp)+1)

    series = Series(y_values, x_values, title_from_data = False)
    chart.series.append(series)
    chart.legend = None

    sheet_delta_V_dot_sim.add_chart(chart, "A13")

    # Mass
    chart = ScatterChart()
    chart.title = "Massenstromverluste über das Netz"
    chart.style = 13
    chart.x_axis.title = "Zeit"
    chart.x_axis.number_format = "hh:mm:ss"
    chart.y_axis.title = "Massenstromverlust [%]"
    chart.width = 30

    x_values = Reference(sheet_delta_V_dot_sim, min_col = 1, min_row = 2, \
                        max_row = len(var_sim.time_stamp)+1)
    y_values = Reference(sheet_delta_V_dot_sim, min_col = x_column+3, \
                        min_row = 2, max_row = len(var_sim.time_stamp)+1)

    series = Series(y_values, x_values, title_from_data = False)
    chart.series.append(series)
    chart.legend = None

    sheet_delta_V_dot_sim.add_chart(chart, "A28")

    fileXLSX.save(fileXLSX_name)


def balance_for_close_vol_flow_dummy(node, var_sim, cntr):
    """Balances volume flow for the gapfilling node.

    :param node: Contains information about the nodes
    :type node: node obj.
    :param var_sim: Variables concerning the simulation
    :type var_sim: var_sim obj.
    :param cntr: Counter
    :type cntr: int
    """

###############################################################################
# BALANCING ###################################################################
###############################################################################

    # INITIALIZATION ##########################################################
    delta_V_dot_sim = []

    V_dot_sum_int, V_dot_feed_sum_int, m_dot_sum_int, m_dot_feed_sum_int \
        = (0 for i in range(4))

    for x_node in range (0, node.nbr_orig_arabic.shape[0]):

        # SUMMATION OF CONSUMERS ##########################################
        if not isinstance(node.V_dot_sim[x_node], type(None)):

            # Summation of measurements
            V_dot_sum_int = V_dot_sum_int+node.V_dot_sim[x_node][cntr]
            if node.temp_flow[x_node][cntr] == None or \
                node.temp_ret[x_node][cntr] == None:
                m_dot_sum_int = m_dot_sum_int+\
                    node.V_dot_sim[x_node][cntr]*10**(-3)*\
                    fcns_phy.H2O_density(50)
            else:
                m_dot_sum_int = m_dot_sum_int+\
                    node.V_dot_sim[x_node][cntr]*10**(-3)*\
                    fcns_phy.H2O_density(mean([node.\
                    temp_flow[x_node][cntr], node.\
                    temp_ret[x_node][cntr]]))

        # SUMMATION OF FEEDERS ############################################
        if not isinstance(node.V_dot_feed[x_node], type(None)):
            V_dot_feed_sum_int = V_dot_feed_sum_int+\
                node.V_dot_feed[x_node][cntr]
            m_dot_feed_sum_int = m_dot_feed_sum_int+\
                node.V_dot_feed[x_node][cntr]*10**(-3)*\
                fcns_phy.H2O_density(mean([node.\
                temp_flow_feed[x_node][cntr], node.\
                temp_ret_feed[x_node][cntr]]))

    # BALANCING ###########################################################
    delta_V_dot_sim = V_dot_feed_sum_int-V_dot_sum_int

    return delta_V_dot_sim