# -*- coding: utf-8 -*-
"""
Author: 4wardEnergy Research GmbH
Date: 2024-05-14
Version: 1.0

Saves the results of the network simulation into an Excel file. The data is organized 
into different sheets within the Excel file for easy analysis and visualization.

Functions:
- Save_Excel: Main function to save the simulation results to an Excel file.
    - Inserts headers and populates data for consumers, feeders, and network losses.
    - Formats the Excel sheets and creates scatter charts for visual representation of data.
"""
"""
Copyright (C) 2024  4wardEnergy Research GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from openpyxl.styles import Font
import numpy as np
from openpyxl.chart import ScatterChart, Reference, Series
from datetime import timedelta
import math

def Save_Excel(var_H2O, var_sim, var_misc, var_unused, line, node, fileXLSX, fileXLSX_name):
    """Saves the results of the network simulation in an excel file.

    :param var_H2O: Variables concerning water
    :type var_H2O: var_H2O obj.

    :param var_sim: Variables concerning the simulation
    :type var_sim: var_sim obj.

    :param var_misc: Miscellaneous variables
    :type var_misc: var_misc obj.

    :param var_unused: Unused variables
    :type var_unused: var_unused obj.

    :param line: Contains information about the pipes
    :type line: line obj.

    :param node: Contains information about the nodes
    :type node: node obj.

    :param fileXLSX: Active excel workbook
    :type fileXLSX: openpyxl.workbook.workbook obj.

    :param fileXLSX_name: Path of the active workbook
    :type fileXLSX_name: str
    """    
    
    Q_dot_feed_sim = []
    
    # print("Speichern")
    ###########################################################################
    ###########################################################################
    # SAVE RESULTS OF NETWORK SIMULATION TO EXCEL WORKSHEET ###################
    ###########################################################################
    ###########################################################################

    ###########################################################################
    # CONSUMERS ###############################################################
    ###########################################################################

    for x_node in range(0, node.nbr_orig_roman.shape[0]):
        Q_dot_cons_sim_add = []
        if node.cons[x_node] != None:
            if node.cons[x_node] == "keine ID":
                sheet_cons = fileXLSX["C_"+str(node.nbr_orig_roman[x_node])]
            else:
                sheet_cons = fileXLSX["C_"+str(node.cons[x_node])]

            # INSERT HEADER
            sheet_cons_caption = ["V\u0307_calc [l/s]", "Q\u0307_calc [kW]", \
                                  "t_VL_calc [°C]", "t_RL_calc [°C]", \
                                  "p_VL_calc [Pa]", "p_RL_calc [Pa]", \
                                  "\u0394t_VL_calc_meas [K]", \
                                  "\u0394t_RL_calc_meas [K]", \
                                  "\u0394p_VL_calc_meas [Pa]", \
                                  "\u0394p_RL_calc_meas [Pa]"]
            for x in range (0, len(sheet_cons_caption)):
                sheet_cons.cell(row = 1, column = x+10).value = \
                    sheet_cons_caption[x]

            for cntr_time in range (0, var_sim.cntr_time_hyd):

                if cntr_time == 0:
                    xx = 0
                else:
                    xx = int(cntr_time*var_sim.delta_time_hyd*60/var_sim.delta_time_therm)

                # CREATE ARRAY
                Q_dot_cons_sim_add = np.append(Q_dot_cons_sim_add, np.array(node.V_dot_sim[x_node][cntr_time]*(node.t_forerun_trans[x_node][xx][0]-node.t_return_trans[x_node][xx][0])*var_H2O.c_p*10**(-3)))

                # FILL WORKSHEET
                sheet_cons.cell(row = cntr_time+2, column = 10).value = node.V_dot_sim[x_node][cntr_time]
                sheet_cons.cell(row = cntr_time+2, column = 11).value = node.Q_dot_sim[x_node][cntr_time]
                sheet_cons.cell(row = cntr_time+2, column = 12).value = node.t_forerun_trans[x_node][xx][0]
                sheet_cons.cell(row = cntr_time+2, column = 13).value = node.t_return_trans[x_node][xx][0]
                sheet_cons.cell(row = cntr_time+2, column = 14).value = node.p_forerun_trans[cntr_time][x_node]
                sheet_cons.cell(row = cntr_time+2, column = 15).value = node.p_return_trans[cntr_time][x_node]
                
                # CALCULATION OF DEVIATIONS
                if isinstance(node.temp_flow[x_node], type(None)):
                    sheet_cons.cell(row = cntr_time+2, column = 16).value = node.temp_flow[x_node][cntr_time]-node.t_forerun_trans[x_node][xx][0]
                else:
                    sheet_cons.cell(row = cntr_time+2, column = 16).value = "-"
                if isinstance(node.temp_ret[x_node], type(None)):
                    sheet_cons.cell(row = cntr_time+2, column = 17).value = node.temp_ret[x_node][cntr_time]-node.t_return_trans[x_node][xx][0]
                else:
                    sheet_cons.cell(row = cntr_time+2, column = 17).value = "-"
                
                if isinstance(node.p_flow_feed[x_node], type(None)):
                    sheet_cons.cell(row = cntr_time+2, column = 18).value = "-"
                else:
                    sheet_cons.cell(row = cntr_time+2, column = 18).value = node.p_flow_feed[x_node][cntr_time]-node.p_forerun_trans[cntr_time][x_node]
                if isinstance(node.p_ret_feed[x_node], type(None)):
                    sheet_cons.cell(row = cntr_time+2, column = 19).value = "-"
                else:
                    sheet_cons.cell(row = cntr_time+2, column = 19).value = node.p_ret_feed[x_node][cntr_time]-node.p_return_trans[cntr_time][x_node]

            node.Q_dot_sim.append(Q_dot_cons_sim_add)

            # FORMATTING
            bold_font = Font(bold = True)
            for cell in sheet_cons["1:1"]:
                cell.font = bold_font
            '''
            # DIAGRAM
            rows = [["Zeit", "delta_t_VL_calc_meas [K]", "delta_t_RL_calc_meas [K]", "delta_p_VL_calc_meas [Pa]", "delta_p_RL_calc_meas [Pa]"]]
            for cntr_time in range (0, var_sim.cntr_time_hyd):
                rows_add = []
                rows_add.append(var_unused.time_excel_sum[cntr_time])
                rows_add.append(var_unused.delta_Q_dot_percent[cntr_time])
                rows.append(rows_add)
        
            for row in rows:
                #var_unused.sheet_delta_Q_dot.append(row)
                sheet_cons.append(row)
        
            chart = ScatterChart()
            chart.title = "Wärmeverluste über das Netz"
            chart.style = 13
            chart.x_axis.title = 'Zeit'
            chart.x_axis.number_format = "hh:mm:ss"
            chart.y_axis.title = 'Wärmeverlust [%]'
            chart.width = 30
        
            xvalues = Reference(var_unused.sheet_delta_Q_dot, min_col = 1, min_row = len(var_unused.time_excel_sum)+3, max_row = 2+2*len(var_unused.time_excel_sum))
            yvalues = Reference(var_unused.sheet_delta_Q_dot, min_col = 2, min_row = len(var_unused.time_excel_sum)+3, max_row = 2+2*len(var_unused.time_excel_sum))
            series = Series(yvalues, xvalues, title_from_data = False)
            chart.series.append(series)
            chart.legend = None
        
            #var_unused.sheet_delta_Q_dot.add_chart(chart, "A13")
            sheet_cons.add_chart(chart, "A13")
            '''
        else:
            node.Q_dot_sim.append([])
    ###########################################################################
    # FEEDERS #################################################################
    ###########################################################################

    for x_node in range(0, node.nbr_orig_roman.shape[0]):
        Q_dot_feed_sim_add = []
        if node.feed_in[x_node] != None:
            sheet_feed = fileXLSX["F_"+str(node.feed_in[x_node])]

            # INSERT HEADER
            sheet_feed_caption = ["V\u0307_calc [l/s]", "Q\u0307_calc [kW]",\
                                  "t_VL_calc [°C]", "t_RL_calc [°C]", \
                                  "p_VL_calc [Pa]", "p_RL_calc [Pa]"]
            for x in range (0, len(sheet_feed_caption)):
                sheet_feed.cell(row = 1, column = x+10).value = \
                    sheet_feed_caption[x]

            for cntr_time in range (0, var_sim.cntr_time_hyd):
                if cntr_time == 0:
                    xx = 0
                else:
                    xx = int(cntr_time*var_sim.delta_time_hyd*60/var_sim.delta_time_therm)

                # CREATE ARRAY
                Q_dot_feed_sim_add = np.append(Q_dot_feed_sim_add, np.array(node.V_dot_feed[x_node][cntr_time]*(node.t_forerun_trans[x_node][xx][0]-node.t_return_trans[x_node][xx][0])*var_H2O.c_p*10**(-3)))
                
                # FILL WORKSHEET
                sheet_feed.cell(row = cntr_time+2, column = 10).value = node.V_dot_feed[x_node][cntr_time]
                sheet_feed.cell(row = cntr_time+2, column = 11).value = node.V_dot_feed[x_node][cntr_time]*(node.t_forerun_trans[x_node][xx][0]-node.t_return_trans[x_node][xx][0])*var_H2O.c_p*10**(-3)
                sheet_feed.cell(row = cntr_time+2, column = 12).value = node.t_forerun_trans[x_node][xx][0]
                sheet_feed.cell(row = cntr_time+2, column = 13).value = node.t_return_trans[x_node][xx][0]
                sheet_feed.cell(row = cntr_time+2, column = 14).value = node.p_forerun_trans[cntr_time][x_node]
                sheet_feed.cell(row = cntr_time+2, column = 15).value = node.p_return_trans[cntr_time][x_node]

            Q_dot_feed_sim.append(Q_dot_feed_sim_add)
        else:
            Q_dot_feed_sim.append(None)

            # FORMATTING
            bold_font = Font(bold = True)
            for cell in sheet_feed["1:1"]:
                cell.font = bold_font

    ###########################################################################
    # BALANCING OF THE SIMULATED POWER ########################################
    ###########################################################################

    # BALANCING
    delta_Q_dot_sim, delta_Q_dot_sim_percent = [], []
    for cntr_time in range (0, var_sim.cntr):
        Q_dot_feed_sim_sum = 0
        Q_dot_cons_sim_sum = 0
        for x_node in range (0, node.nbr_matrix.shape[0]):
            try:
                if Q_dot_feed_sim[x_node] == None:
                    pass
            except:
                Q_dot_feed_sim_sum = Q_dot_feed_sim_sum+Q_dot_feed_sim[x_node][cntr_time]
            if not isinstance(node.Q_dot_sim[x_node], type(None)):
                Q_dot_cons_sim_sum = Q_dot_cons_sim_sum+node.Q_dot_sim[x_node][cntr_time]
        delta_Q_dot_sim.append(Q_dot_feed_sim_sum-Q_dot_cons_sim_sum)
        if Q_dot_feed_sim_sum == 0:
            delta_Q_dot_sim_percent.append("-")
        else:
            delta_Q_dot_sim_percent.append(delta_Q_dot_sim[-1]/Q_dot_feed_sim_sum*100)

    # CREATE NEW SHEET
    sheet_name_Q_dot_balance_sim = "Q_dot_Bilanz_sim"
    try:
        del fileXLSX[sheet_name_Q_dot_balance_sim]
    except:
        pass
    fileXLSX.create_sheet(sheet_name_Q_dot_balance_sim)
    fileXLSX.save(fileXLSX_name)

    sheet_delta_Q_dot_sim = fileXLSX[sheet_name_Q_dot_balance_sim]

    # TIME STAMPS
    sheet_delta_Q_dot_sim.cell(row = 1, column = 1).value = "Zeit ↓ Nr. → Q\u0307 [kW] ↘"
    time_excel = var_sim.time_sim_start
    var_unused.time_excel_sum = []
    row_excel = 2
    while time_excel <= var_sim.time_sim:
        sheet_delta_Q_dot_sim.cell(row = row_excel, column = 1).value = time_excel
        sheet_delta_Q_dot_sim.cell(row = row_excel, column = 1).style = var_misc.date_style
        var_unused.time_excel_sum.append(time_excel)
        time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
        row_excel += 1

    # POWERS
    x_column = 2
    for x_node in range (0, node.nbr_matrix.shape[0]):
        time_excel = var_sim.time_sim_start
        if node.cons[x_node] != None:
            if node.cons[x_node] == "keine ID":
                sheet_delta_Q_dot_sim.cell(row = 1, column = x_column).value = node.nbr_orig_roman[x_node]
                row_excel = 2
                while time_excel <= var_sim.time_sim:
                #while time_excel <= var_sim.time_sim_end:
                    sheet_delta_Q_dot_sim.cell(row = row_excel, column = x_column).value = "-"
                    row_excel += 1
                    time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
            else:
                sheet_delta_Q_dot_sim.cell(row = 1, column = x_column).value = node.cons[x_node]
                cntr_excel = 0
                row_excel = 2
                while time_excel <= var_sim.time_sim:
                #while time_excel <= var_sim.time_sim_end:
                    sheet_delta_Q_dot_sim.cell(row = row_excel, column = x_column).value = node.Q_dot_sim[x_node][cntr_excel]
                    cntr_excel += 1
                    row_excel += 1
                    time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
            x_column += 1
        else:
            if node.feed_in[x_node] != None:
                sheet_delta_Q_dot_sim.cell(row = 1, column = x_column).value = "F_"+str(node.feed_in[x_node])
                cntr_excel = 0
                row_excel = 2
                while time_excel <= var_sim.time_sim:
                #while time_excel <= var_sim.time_sim_end:
                    sheet_delta_Q_dot_sim.cell(row = row_excel, column = x_column).value = Q_dot_feed_sim[x_node][cntr_excel]
                    cntr_excel += 1
                    row_excel += 1
                    time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
                x_column += 1
    sheet_delta_Q_dot_sim.cell(row = 1, column = x_column).value = "\u0394Q\u0307_sim [kW]"
    sheet_delta_Q_dot_sim.cell(row = 1, column = x_column+1).value = "\u0394Q\u0307_sim [%]"
    sheet_delta_Q_dot_sim.cell(row = 1, column = x_column+2).value = "t_Boden [°C]"
    row_excel = 2
    for x_row in range (0, len(delta_Q_dot_sim)):
        sheet_delta_Q_dot_sim.cell(row = row_excel, column = x_column).value = delta_Q_dot_sim[x_row]
        sheet_delta_Q_dot_sim.cell(row = row_excel, column = x_column+1).value = delta_Q_dot_sim_percent[x_row]
        sheet_delta_Q_dot_sim.cell(row = row_excel, column = x_column+2).value = var_sim.temp_soil[x_row]
        row_excel += 1

    # FORMATTING
    bold_font = Font(bold = True)
    for cell in sheet_delta_Q_dot_sim["1:1"]:
        cell.font = bold_font
    sheet_delta_Q_dot_sim.freeze_panes = 'A2'
    sheet_delta_Q_dot_sim.column_dimensions["A"].width = 25

    # DIAGRAM
    rows = [["Zeit", "\u0394Q\u0307_sim [%]"]]
    #print(delta_Q_dot_sim_percent)
    for cntr_time in range (0, len(var_unused.time_excel_sum)):
        rows_add = []
        rows_add.append(var_unused.time_excel_sum[cntr_time])
        rows_add.append(delta_Q_dot_sim_percent[cntr_time])
        rows.append(rows_add)
    
    for row in rows:
        sheet_delta_Q_dot_sim.append(row)

    chart = ScatterChart()
    chart.title = "Wärmeverluste über das Netz"
    chart.style = 13
    chart.x_axis.title = 'Zeit'
    chart.x_axis.number_format = "hh:mm:ss"
    chart.y_axis.title = 'Wärmeverlust [%]'
    chart.width = 30
    
    xvalues = Reference(sheet_delta_Q_dot_sim, min_col = 1, min_row = len(var_unused.time_excel_sum)+3, max_row = 2+2*len(var_unused.time_excel_sum))
    yvalues = Reference(sheet_delta_Q_dot_sim, min_col = 2, min_row = len(var_unused.time_excel_sum)+3, max_row = 2+2*len(var_unused.time_excel_sum))
    series = Series(yvalues, xvalues, title_from_data = False)
    chart.series.append(series)
    chart.legend = None
    
    sheet_delta_Q_dot_sim.add_chart(chart, "A13")

    fileXLSX.save(fileXLSX_name)

    ###########################################################################
    # NETWORK LOSSES - FLOW ###################################################
    ###########################################################################

    # CREATE NEW SHEET
    sheet_name_gride_loss_forerun = "NV_VL_sim"
    try:
        del fileXLSX[sheet_name_gride_loss_forerun]
    except:
        pass
    fileXLSX.create_sheet(sheet_name_gride_loss_forerun)
    fileXLSX.save(fileXLSX_name)

    sheet_grid_loss_forerun = fileXLSX[sheet_name_gride_loss_forerun]

    # NAME COLUMNS
    sheet_grid_loss_forerun.cell(row = 1, column = 1).value = "Zeit ↓ Nr. → Q\u0307 [kW] ↘"
    for x_line in range(0, len(line.nbr_orig)):
        sheet_grid_loss_forerun.cell(row = 1, column = x_line+2).value = int(line.nbr_orig[x_line])

    # INSERT TIME STAMPS AND VALUES
    time_excel = var_sim.time_sim_start
    row_excel = 2
    var_sim.cntr = 0
    while time_excel <= var_sim.time_sim:
    #while time_excel <= var_sim.time_sim_end:
        sheet_grid_loss_forerun.cell(row = row_excel, column = 1).value = time_excel
        sheet_grid_loss_forerun.cell(row = row_excel, column = 1).style = var_misc.date_style
        for x_line in range(0, len(line.nbr_orig)):
            sheet_grid_loss_forerun.cell(row = row_excel, column = x_line+2).value = line.Q_dot_forerun_line_loss[var_sim.cntr][x_line]
        time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
        row_excel += 1
        var_sim.cntr += 1

    # FORMATTING
    bold_font = Font(bold = True)
    for cell in sheet_grid_loss_forerun["1:1"]:
        cell.font = bold_font
    sheet_grid_loss_forerun.column_dimensions["A"].width = 25

    ###########################################################################
    # NETWORK LOSSES - RETURN #################################################
    ###########################################################################

    # CREATE NEW SHEET
    sheet_name_gride_loss_return = "NV_RL_sim"
    try:
        del fileXLSX[sheet_name_gride_loss_return]
    except:
        pass
    fileXLSX.create_sheet(sheet_name_gride_loss_return)
    fileXLSX.save(fileXLSX_name)

    sheet_grid_loss_return = fileXLSX[sheet_name_gride_loss_return]

    # NAME COLUMNS
    sheet_grid_loss_return.cell(row = 1, column = 1).value = "Zeit ↓ Nr. → Q\u0307 [kW] ↘"
    for x_line in range(0, len(line.nbr_orig)):
        sheet_grid_loss_return.cell(row = 1, column = x_line+2).value = int(line.nbr_orig[x_line])

    # INSERT TIME STAMPS AND VALUES
    time_excel = var_sim.time_sim_start
    row_excel = 2
    var_sim.cntr = 0
    while time_excel <= var_sim.time_sim:
    #while time_excel <= var_sim.time_sim_end:
        sheet_grid_loss_return.cell(row = row_excel, column = 1).value = time_excel
        sheet_grid_loss_return.cell(row = row_excel, column = 1).style = var_misc.date_style
        for x_line in range(0, len(line.nbr_orig)):
            sheet_grid_loss_return.cell(row = row_excel, column = x_line+2).value = line.Q_dot_return_line_loss[var_sim.cntr][x_line]
        time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
        row_excel += 1
        var_sim.cntr += 1

    # FORMATTING
    bold_font = Font(bold = True)
    for cell in sheet_grid_loss_return["1:1"]:
        cell.font = bold_font
    sheet_grid_loss_return.column_dimensions["A"].width = 25

    ###########################################################################
    # NETWORK LOSSES - FLOW - RELATIVE ########################################
    ###########################################################################
    ''' NOTE: This part is not working yet. The problem is that the relative
    values are not calculated correctly. The values are not relative to the
    total heat loss of the network but to the heat loss of the line. This
    means that the sum of the relative values is not 100% but 100%*number of
    lines. This is not a problem for the absolute values because the sum of
    the absolute values is equal to the total heat loss of the network.

    # CREATE NEW SHEET
    sheet_name_gride_loss_forerun_rel = "NV_VL_sim_rel"
    try:
        del fileXLSX[sheet_name_gride_loss_forerun_rel]
    except:
        pass
    fileXLSX.create_sheet(sheet_name_gride_loss_forerun_rel)
    fileXLSX.save(file_input)

    sheet_grid_loss_forerun_rel = fileXLSX[sheet_name_gride_loss_forerun_rel]

    # NAME COLUMNS
    sheet_grid_loss_forerun_rel.cell(row = 1, column = 1).value = "Zeit ↓ Nr. →  [W/(m·W)] ↘"
    for x_line in range(0, len(line.nbr_orig)):
        sheet_grid_loss_forerun_rel.cell(row = 1, column = x_line+2).value = int(line.nbr_orig[x_line])

    # INSERT TIME STAMPS AND VALUES
    time_excel = var_sim.time_sim_start
    row_excel = 2
    var_sim.cntr = 0
    while time_excel <= var_sim.time_sim:
        sheet_grid_loss_forerun_rel.cell(row = row_excel, column = 1).value = time_excel
        sheet_grid_loss_forerun_rel.cell(row = row_excel, column = 1).style = var_misc.date_style
        for x_line in range(0, len(line.nbr_orig)):
            sheet_grid_loss_forerun_rel.cell(row = row_excel, column = x_line+2).value = line.q_dot_forerun_line_loss[var_sim.cntr][x_line]
        time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
        row_excel += 1
        var_sim.cntr += 1

    # FORMATTING
    bold_font = Font(bold = True)
    for cell in sheet_grid_loss_forerun_rel["1:1"]:
        cell.font = bold_font
    sheet_grid_loss_forerun_rel.column_dimensions["A"].width = 25
    
    ###########################################################################
    # NETWORK LOSSES - RETURN - RELATIVE ######################################
    ###########################################################################

    # CREATE NEW SHEET
    sheet_name_gride_loss_return_rel = "NV_RL_sim_rel"
    try:
        del fileXLSX[sheet_name_gride_loss_return_rel]
    except:
        pass
    fileXLSX.create_sheet(sheet_name_gride_loss_return_rel)
    fileXLSX.save(file_input)

    sheet_grid_loss_return_rel = fileXLSX[sheet_name_gride_loss_return_rel]

    # NAME COLUMNS
    sheet_grid_loss_return_rel.cell(row = 1, column = 1).value = "Zeit ↓ Nr. →  [W/(m·W)] ↘"
    for x_line in range(0, len(line.nbr_orig)):
        sheet_grid_loss_return_rel.cell(row = 1, column = x_line+2).value = int(line.nbr_orig[x_line])

    # INSERT TIME STAMPS AND VALUES
    time_excel = var_sim.time_sim_start
    row_excel = 2
    var_sim.cntr = 0
    while time_excel <= var_sim.time_sim:
        sheet_grid_loss_return_rel.cell(row = row_excel, column = 1).value = time_excel
        sheet_grid_loss_return_rel.cell(row = row_excel, column = 1).style = var_misc.date_style
        for x_line in range(0, len(line.nbr_orig)):
            sheet_grid_loss_return_rel.cell(row = row_excel, column = x_line+2).value = line.q_dot_return_line_loss[var_sim.cntr][x_line]
        time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
        row_excel += 1
        var_sim.cntr += 1

    # FORMATTING
    bold_font = Font(bold = True)
    for cell in sheet_grid_loss_return_rel["1:1"]:
        cell.font = bold_font
    sheet_grid_loss_return_rel.column_dimensions["A"].width = 25
    '''
    ###########################################################################
    # FLOW RATES - FLOW #######################################################
    ###########################################################################

    # CREATE NEW SHEET
    sheet_name_flow_rate_forerun = "v_VL_m"
    try:
        del fileXLSX[sheet_name_flow_rate_forerun]
    except:
        pass
    fileXLSX.create_sheet(sheet_name_flow_rate_forerun)
    fileXLSX.save(fileXLSX_name)

    sheet_flow_rate_forerun = fileXLSX[sheet_name_flow_rate_forerun]

    # NAME COLUMNS
    sheet_flow_rate_forerun.cell(row = 1, column = 1).value = "Zeit ↓ Nr. → v [m/s] ↘"
    for x_line in range(0, len(line.nbr_orig)):
        sheet_flow_rate_forerun.cell(row = 1, column = x_line+2).value = int(line.nbr_orig[x_line])

    # INSERT TIME STAMPS AND VALUES
    time_excel = var_sim.time_sim_start
    row_excel = 2
    var_sim.cntr = 0
    while time_excel <= var_sim.time_sim:
    #while time_excel <= var_sim.time_sim_end:
        sheet_flow_rate_forerun.cell(row = row_excel, column = 1).value = time_excel
        sheet_flow_rate_forerun.cell(row = row_excel, column = 1).style = var_misc.date_style
        for x_line in range(0, len(line.nbr_orig)):
            sheet_flow_rate_forerun.cell(row = row_excel, column = x_line+2).value = abs(line.m_int_forerun_trans[x_line][var_sim.cntr][0])/var_H2O.rho/(line.dia[x_line]**2*math.pi/4) 
        time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
        row_excel += 1
        var_sim.cntr += 1

    # FORMATTING
    bold_font = Font(bold = True)
    for cell in sheet_flow_rate_forerun["1:1"]:
        cell.font = bold_font
    sheet_flow_rate_forerun.column_dimensions["A"].width = 20
    
    ###########################################################################
    # FLOW RATES - RETURN #####################################################
    ###########################################################################

    # CREATE NEW SHEET
    sheet_name_flow_rate_return = "v_RL_m"
    try:
        del fileXLSX[sheet_name_flow_rate_return]
    except:
        pass
    fileXLSX.create_sheet(sheet_name_flow_rate_return)
    fileXLSX.save(fileXLSX_name)

    sheet_flow_rate_return = fileXLSX[sheet_name_flow_rate_return]

    # NAME COLUMNS
    sheet_flow_rate_return.cell(row = 1, column = 1).value = "Zeit ↓ Nr. → v [m/s] ↘"
    for x_line in range(0, len(line.nbr_orig)):
        sheet_flow_rate_return.cell(row = 1, column = x_line+2).value = int(line.nbr_orig[x_line])

    # INSERT TIME STAMPS AND VALUES
    time_excel = var_sim.time_sim_start
    row_excel = 2
    var_sim.cntr = 0
    while time_excel <= var_sim.time_sim:
    #while time_excel <= var_sim.time_sim_end:
        sheet_flow_rate_return.cell(row = row_excel, column = 1).value = time_excel
        sheet_flow_rate_return.cell(row = row_excel, column = 1).style = var_misc.date_style
        for x_line in range(0, len(line.nbr_orig)):
            sheet_flow_rate_return.cell(row = row_excel, column = x_line+2).value = abs(line.m_int_return_trans[x_line][var_sim.cntr][0])/var_H2O.rho/(line.dia[x_line]**2*math.pi/4)
        time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
        row_excel += 1
        var_sim.cntr += 1

    # FORMATTING
    bold_font = Font(bold = True)
    for cell in sheet_flow_rate_return["1:1"]:
        cell.font = bold_font
    sheet_flow_rate_return.column_dimensions["A"].width = 20

    #sheet_flow_rate_return.conditional_format('B2:H14', {'type': '3_color_scale'})

    fileXLSX.save(fileXLSX_name)


    ###########################################################################
    # NETWORK WORST POINT #####################################################
    ###########################################################################

    # CREATE NEW SHEET
    sheet_name_netzschlechtpunkt = "Netzschlechtpunkt"
    try:
        del fileXLSX[sheet_name_netzschlechtpunkt]
    except:
        pass
    fileXLSX.create_sheet(sheet_name_netzschlechtpunkt)
    fileXLSX.save(fileXLSX_name)

    sheet_netzschlechtpunkt = fileXLSX[sheet_name_netzschlechtpunkt]

    # NAME COLUMNS
    sheet_netzschlechtpunkt.cell(row = 1, column = 1).value = "Zeitschritt"
    sheet_netzschlechtpunkt.cell(row = 1, column = 2).value = "Knoten Nr."
    sheet_netzschlechtpunkt.cell(row = 1, column = 3).value = "Abnehmer ID"
    sheet_netzschlechtpunkt.cell(row = 1, column = 4).value = "Δp_calc [Pa]"

    # ITERATE THROUGH TIME STEPS
    time_excel = var_sim.time_sim_start
    cntr_time = 0
    row_excel = 2
    while time_excel <= var_sim.time_sim:
        list_p_forerun = []
        list_p_return = []
        for x_node in range (0, node.nbr_matrix.shape[0]):
            # List all forerun and return pressures
            list_p_forerun.append(node.p_forerun_trans[cntr_time][x_node])
            list_p_return.append(node.p_return_trans[cntr_time][x_node])
        # Calculate differential pressure
        list_p_diff = list(np.array(list_p_forerun)-np.array(list_p_return))
        # Get index of node with lowest pressure
        min_p_diff = min(list_p_diff)
        x_node_min = list_p_diff.index(min_p_diff)
        # Write to excel
        sheet_netzschlechtpunkt.cell(row = row_excel, column = 1).value = time_excel
        sheet_netzschlechtpunkt.cell(row = row_excel, column = 1).style = var_misc.date_style
        sheet_netzschlechtpunkt.cell(row = row_excel, column = 2).value = node.nbr_orig_roman[x_node_min]
        sheet_netzschlechtpunkt.cell(row = row_excel, column = 3).value = node.cons[x_node_min]
        sheet_netzschlechtpunkt.cell(row = row_excel, column = 4).value = min_p_diff

        # Next time step
        time_excel = time_excel+timedelta(seconds = 60*var_sim.delta_time_hyd)
        row_excel += 1
        cntr_time += 1



    # FORMATTING
    bold_font = Font(bold = True)
    for cell in sheet_netzschlechtpunkt["1:1"]:
        cell.font = bold_font
    sheet_netzschlechtpunkt.column_dimensions["A"].width = 20

    fileXLSX.save(fileXLSX_name)
    
    '''
    workbook1 = xlsxwriter.Workbook(file_input)
    existingWorksheet = workbook1.get_worksheet_by_name("v-RL")
    #existingWorksheet.write_row(0,0,'xyz')
    existingWorksheet.write(0, 0, 'Hello Excel')
    #existingWorksheet.conditional_format('B2:Z12', {'type': 'data_bar'})
    workbook1.close()'''