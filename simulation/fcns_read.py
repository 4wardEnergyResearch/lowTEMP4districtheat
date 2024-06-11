# -*- coding: utf-8 -*-
"""
Author: 4wardEnergy Research GmbH
Date: 2024-05-14
Version: 1.0

Functions to read and process data from the topology file and feeder/consumer 
CSV files.

Functions:
- read_data: Wrapper function for the other functions in this module. Reads the data from the topology file.
- read_grid_setup: Reads the grid setup from the topology file
- read_cons: Reads consumer data from the time series CSVs.
- read_cons_non_ID: Fills the node object with blanks for the case of a consumer without ID.
- read_feed: Reads feeder data from the time series CSVs.
"""
"""
Copyright (C) 2024  4wardEnergy Research GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import numpy as np

import os
import csv
import openpyxl
import shutil
from openpyxl.styles import Font
from datetime import datetime, timedelta

from options import *
from simulation import fcns_read
from simulation import Auxiliary_functions

def read_data(var_misc, var_sim, var_gaps, file_input):
    """Reads the data from the excel file located at file_input.

    :param var_misc: Miscellaneous variables
    :type var_misc: var_misc obj.

    :param var_sim: Variables concerning the simulation
    :type var_sim: var_sim obj.

    :param var_gaps: Variables concerning the gapfilling
    :type var_gaps: var_gaps obj.

    :param file_input: Path of the input Excel file
    :type file_input: str

    :return: active excel workbook.
    :rtype: openpyxl obj.

    :return: Path of the active workbook
    :rtype: str

    :return: Contains information about the nodes
    :rtype: node obj.

    :return: Contains information about the pipes
    :rtype: line obj.
    """    

###############################################################################
###############################################################################
# CREATE EXCEL FILE ###########################################################
###############################################################################
###############################################################################

    # Set name of the new excel file
    fileXLSX_name = file_input[:-(len(file_input.split(".")[-1])+1)]+\
        "_"+str(datetime.now().strftime(format = "%Y-%m-%d_%H-%M"))+".xlsx"
    # Copy original Excel file to new Excel file
    shutil.copy(file_input, fileXLSX_name)
    # Open new excel file
    fileXLSX = openpyxl.load_workbook (fileXLSX_name)

###############################################################################
###############################################################################
# READ NETWORK DATA ###########################################################
###############################################################################
###############################################################################

    node, line = fcns_read.read_grid_setup(fileXLSX, fileXLSX_name)

###############################################################################
###############################################################################
# READ CONSUMER MEASUREMENTS ##################################################
###############################################################################
###############################################################################

    # INITIALIZATION ##########################################################
    node.V_dot, node.temp_flow, node.temp_ret, node.Q_dot, \
        node.p_flow, node.p_ret, node.error, node.csv_exist, node.t_out = ([] for i in range(9))
    '''
    node.error: "OKAY_m" = okay, internal error recognition
                "FEHLER_m" = error, internal error recognition
                "OKAY_a" = okay, external error recognition
                "FEHLER_a" = error, external error recognition
                "FEHLER_l" = error, empty line in .csv file
                "FEHLER" = error, consumer ID or .csv file missing
    '''

    for x_node in range(0, node.nbr_matrix.shape[0]):
        if node.cons[x_node] != None:
            ###################################################################
            # CONSUMERS WITHOUT ID ############################################
            ###################################################################
            # Create new sheet, fill values with ""
            if node.cons[x_node] == "keine ID":

                # CSV FILE EXISTENCE INFO #####################################
                node.csv_exist = np.append(node.csv_exist, "")

                # READ ########################################################
                fileXLSX, node = fcns_read.read_cons_non_ID(fileXLSX, \
                    fileXLSX_name, x_node, var_sim, var_misc, node)

            ###################################################################
            # CONSUMERS WITH ID ###############################################
            ###################################################################
            else:
                # READ ########################################################
                fileXLSX, node, node.csv_exist = fcns_read.read_cons(fileXLSX, \
                    fileXLSX_name, x_node, var_sim, var_misc, var_gaps, node, node.csv_exist)

        #######################################################################
        # NOT A CONSUMER ######################################################
        #######################################################################

        else:
            # CSV FILE EXISTENCE INFO #########################################
            node.csv_exist = np.append(node.csv_exist, "")

            # FILL NODE OBJECT ################################################
            node.V_dot.append(None)
            node.temp_flow.append(None)
            node.temp_ret.append(None)
            node.Q_dot.append(None)
            node.p_flow.append(None)
            node.p_ret.append(None)
            node.error.append(None)

###############################################################################
# CSV FILE EXISTENCE INFO #####################################################
###############################################################################
# If CSV-file exists, the field in column "CSV" is filled with "v."
    sheet_nodes = fileXLSX["Knoten"]
    row_xlsx = 2
    x_node = 0
    for row in sheet_nodes['B2':'B401']:
        for col in row:
            if col.value != None and sheet_nodes.cell(row = row_xlsx, \
                column = 2).value == "ja":
                sheet_nodes.cell(row = row_xlsx, column = 14).value = \
                    node.csv_exist[x_node]
                x_node += 1
            row_xlsx += 1

    fileXLSX.save(fileXLSX_name)

###############################################################################
###############################################################################
# READ FEEDERS DATA ###########################################################
###############################################################################
###############################################################################

    # INITIALIZATION ##########################################################
    node.Q_dot_feed, node.V_dot_feed, node.temp_flow_feed, \
        node.temp_ret_feed, node.p_flow_feed, node.p_ret_feed, \
        node.error_feed = ([] for i in range(7))

    for x_node in range (0, node.nbr_orig_arabic.shape[0]):

    ###########################################################################
    # FEEDERS WITH ID #########################################################
    ###########################################################################
        if node.feed_in[x_node] != None:
            # READ ############################################################
            fileXLSX, node = fcns_read.read_feed(fileXLSX, fileXLSX_name, \
                x_node, var_sim, var_misc, var_gaps, node)

    ###########################################################################
    # FEEDERS WITHOUT ID ######################################################
    ###########################################################################

        else:
            node.Q_dot_feed.append(None)
            node.V_dot_feed.append(None)
            node.temp_flow_feed.append(None)
            node.temp_ret_feed.append(None)
            node.p_flow_feed.append(None)
            node.p_ret_feed.append(None)
            node.error_feed.append(None)

###############################################################################
# SAVE EXCEL FILE #############################################################
###############################################################################
    fileXLSX.save(fileXLSX_name)

###############################################################################
# OUTPUT ######################################################################
###############################################################################
    return fileXLSX, fileXLSX_name, node, line

def read_grid_setup (fileXLSX, fileXLSX_name):
    """Reads the grid setup. Called by read_data.

    :param fileXLSX: Active excel workbook (openpyxl obj.)
    :type fileXLSX: openpyxl obj.

    :param fileXLSX_name: Path of the active workbook (str)
    :type fileXLSX_name: str

    :return node: Contains information about the nodes (node obj.)
    :rtype node: node obj.
    """
###############################################################################
# READ NODES ##################################################################
###############################################################################

    # INITIALIZATION ##########################################################
    class node:
        nbr_orig_roman, nbr_orig_arabic, nbr_matrix, x_coord, y_coord, \
        h_coord, distrib, cons, Q_dot_max, Q_year, H2O_year, cons_type, \
        p_ref, feed_in, hist_data_available, building_type, kWh_m3, gapfilling_node, \
        gapfilling_mode, gapfilling_override, p_offset  = ([] for i in range(21))

    # READ ####################################################################
    # BOOKMARK: Read nodes from topology
    sheet = fileXLSX["Knoten"]
    row_xlsx = 2
    nbr_matrix_add = 0

    for row in sheet['B2':'B401']:
        for col in row:
            if col.value != None and sheet.cell(row = row_xlsx, \
                column = 2).value == "ja":

                # Node number (roman numeral) in excel sheet [-]
                node.nbr_orig_roman = np.append(node.nbr_orig_roman, \
                    np.array([sheet.cell(row = row_xlsx, \
                    column = 1).value]), axis = 0)

                # Node number (arabic numeral) in excel sheet [-]
                node.nbr_orig_arabic = np.append(node.nbr_orig_arabic, \
                    np.array([Auxiliary_functions.romanToInt(sheet.cell(\
                    row = row_xlsx, column = 1).value)]), axis = 0)

                # Node number in coupling matrix [-]
                node.nbr_matrix = np.append(node.nbr_matrix, \
                    np.array([nbr_matrix_add]), axis = 0)
                nbr_matrix_add += 1

                # X-Coordinate [m]
                node.x_coord = np.append(node.x_coord, \
                    np.array([sheet.cell(row = row_xlsx, column = 3).value]), \
                    axis = 0)

                # Y-Coordinate [m]
                node.y_coord = np.append(node.y_coord, \
                    np.array([sheet.cell(row = row_xlsx, column = 4).value]), \
                    axis = 0)

                # Geodetic height [m]
                node.h_coord = np.append(node.h_coord, \
                    np.array([sheet.cell(row = row_xlsx, column = 5).value]), \
                    axis = 0)

                # Type of node [-]
                node.distrib = np.append(node.distrib, \
                    np.array([sheet.cell(row = row_xlsx, column = 6).value]), \
                    axis = 0)
                # "x"...distributor node
                # ""....Not a distributor node

                # ID of consumer [-]
                node.cons = np.append(node.cons, np.array([sheet.cell(row = \
                    row_xlsx, column = 9).value]), axis = 0)

                # Power of heat transfer station [kW]
                node.Q_dot_max = np.append(node.Q_dot_max, \
                    np.array([sheet.cell(row = row_xlsx, \
                    column = 10).value]), axis = 0)

                # Annual heat demand of the heat transfer station [kWh/a]
                node.Q_year = np.append(node.Q_year, \
                    np.array([sheet.cell(row = row_xlsx, \
                    column = 11).value]), axis = 0)

                # Annual water demand of the heat transfer station [m³/a]
                node.H2O_year = np.append(node.H2O_year, \
                    np.array([sheet.cell(row = row_xlsx, \
                    column = 12).value]), axis = 0)

                # Type of building supplied by the heat transfer station [-] (Legacy)
                node.cons_type = np.append(node.cons_type, \
                    np.array([sheet.cell(row = row_xlsx, \
                    column = 13).value]), axis = 0)

                # Node number (w.r.t. coupling matrix) of reference pressure node [-]
                if sheet.cell(row = row_xlsx, column = 7).value == "x":
                    # Check if the reference node is a feeder node:
                    if sheet.cell(row = row_xlsx, column = 8).value == None:
                        raise Exception("Node " + str(node.nbr_orig_roman[-1]) + " is set as the reference point for pressure, but it is not a feeder. Please set the reference point to a feeder.")
                    node.p_ref.append(node.nbr_matrix[-1])

                # Node number (w.r.t. coupling matrix) of feeder(s) [-]
                if sheet.cell(row = row_xlsx, column = 8).value != None:
                    node.feed_in.append(sheet.cell(row = row_xlsx, \
                        column = 8).value)
                else:
                    node.feed_in.append(None)

                # building type [str]
                node.building_type = np.append(node.building_type, np.array([sheet.cell(row = \
                    row_xlsx, column = 17).value]), axis = 0)   

                # Historical data availability [bool]
                node.hist_data_available = np.append(node.hist_data_available, np.array([sheet.cell(row = \
                    row_xlsx, column = 18).value]), axis = 0)

                # kWh/m3 for gapfilling [kWh/m3]
                node.kWh_m3 = np.append(node.kWh_m3, np.array([sheet.cell(row = \
                    row_xlsx, column = 22).value]), axis = 0)

                # Flag for gapfilling node [bool]
                if sheet.cell(row = row_xlsx, column = 23).value == "x":
                    node.gapfilling_node.append(1)
                else:
                    node.gapfilling_node.append(0)

                # Automatically chosen gapfilling mode [str]
                node.gapfilling_mode = np.append(node.gapfilling_mode, np.array([sheet.cell(row = \
                    row_xlsx, column = 24).value]), axis = 0)
                
                # Gapfilling mode override [str]
                node.gapfilling_override = np.append(node.gapfilling_override, np.array([sheet.cell(row = \
                    row_xlsx, column = 25).value]), axis = 0)
                
                # Pressure offset [Pa]
                if sheet.cell(row = row_xlsx, column = 26).value != None:
                    node.p_offset = np.append(node.p_offset, np.array([sheet.cell(row = \
                        row_xlsx, column = 26).value]), axis = 0)
                else:
                    node.p_offset = np.append(node.p_offset, np.array([0]), axis = 0)



            row_xlsx += 1

###############################################################################
# READ LINES ##################################################################
###############################################################################

    class line:

        # INITIALIZATION ######################################################
        nbr_orig, nbr_matrix, node_start, node_end, l, dia, lambd, zeta, \
            htc = ([] for i in range(9))

    # READ ####################################################################
    # BOOKMARK: Read lines from topology
    sheet = fileXLSX["Leitungen"]
    row_xlsx = 2
    line_nbr_matrix = 0

    for row in sheet['B2':'B401']:
        for col in row:
            if col.value != None and sheet.cell(row = row_xlsx, \
                column = 2).value == "ja":

                # Pipe number [-]
                line.nbr_orig = np.append(line.nbr_orig, \
                    np.array([sheet.cell(row = row_xlsx, column = 1).value]), \
                    axis = 0)

                # Pipe number in coupling matrix [-]
                line.nbr_matrix = np.append(line.nbr_matrix, \
                        np.array([line_nbr_matrix]), axis = 0)
                line_nbr_matrix += 1

                for x_node in range(0, node.nbr_matrix.shape[0]):
                    if Auxiliary_functions.romanToInt(sheet.cell(row = \
                        row_xlsx, column = 3).value) == \
                        node.nbr_orig_arabic[x_node]:

                        # Start node number in coupling matrix [-]
                        line.node_start = np.append(line.node_start, \
                            np.array([x_node]), axis = 0)

                    if Auxiliary_functions.romanToInt(sheet.cell(row = \
                        row_xlsx, column = 4).value) == \
                        node.nbr_orig_arabic[x_node]:

                        # End node number in coupling matrix[-]
                        line.node_end = np.append(line.node_end, \
                            np.array([x_node]), axis = 0)

                # Pipe length [m]
                line.l = np.append(line.l, np.array([sheet.cell(row = \
                    row_xlsx, column = 5).value]), axis = 0)

                # Pipe diameter [m]
                line.dia = np.append(line.dia, np.array([sheet.cell(row = \
                    row_xlsx, column = 6).value/1000]), axis = 0)

                # friction coefficient of the pipe [-]
                line.lambd = np.append(line.lambd, np.array([sheet.cell(row = \
                    row_xlsx, column = 7).value]), axis = 0)

                # Sum of drag coefficients of the pipe [-]
                line.zeta = np.append(line.zeta, np.array([sheet.cell(row = \
                    row_xlsx, column = 8).value]), axis = 0)

                # Heat transition coefficient of the pipe [W/mK]
                line.htc = np.append(line.htc, np.array([sheet.cell(row = \
                    row_xlsx, column = 9).value]), axis = 0)

            row_xlsx += 1

###############################################################################
# DELETE UNUSED WORKSHEETS ####################################################
###############################################################################
# All sheets except "Knoten" and "Leitungen"
    sheet_names = fileXLSX.get_sheet_names()
    for x_sheet in range (0, len(sheet_names)):
        if sheet_names[x_sheet] == sheet_names[x_sheet] == "Knoten" or \
            sheet_names[x_sheet] == "Leitungen":
            continue
        else:
            fileXLSX.remove_sheet(fileXLSX.\
                                  get_sheet_by_name(sheet_names[x_sheet]))
    fileXLSX.save(fileXLSX_name)

###############################################################################
# CHECKS ######################################################################
###############################################################################

    # Check if more than one node was designated as gapfilling node
    if sum(node.gapfilling_node) > 1:
        raise Exception("More than one node was designated as gapfilling node. \
                        Please check the input file.")

###############################################################################
# OUTPUT #####################################################################
###############################################################################

    return (node, line)



def read_cons (fileXLSX, fileXLSX_name, x_node, var_sim, var_misc, var_gaps, node, \
               csv_exist):
    """Reads consumer data stored in separate csv files in "Daten" folder.
    Called by read_data.

    :param fileXLSX: Active excel workbook (openpyxl obj.)
    :type fileXLSX: openpyxl obj.

    :param fileXLSX_name: Path of the active workbook (str)
    :type fileXLSX_name: str

    :param x_node: Node number (w.r.t. coupling matrix) of the consumer (int)
    :type x_node: int

    :param var_sim: Variables concerning the simulation (var_sim obj.)
    :type var_sim: var_sim obj.

    :param var_misc: Miscellaneous variables (var_misc obj.)
    :type var_misc: var_misc obj.

    :param var_gaps: Variables concerning the gapfilling
    :type var_gaps: var_gaps obj.

    :param node: Contains information about the nodes (node obj.)
    :type node: node obj.

    :return fileXLSX: Active excel workbook (openpyxl obj.)
    :rtype fileXLSX: openpyxl obj.

    :return node: Contains information about the nodes (node obj.)
    :rtype node: node obj.

    :return csv_exist: "v." if csv file exists, "n.v." if not (str)
    :rtype csv_exist: str

    """

###############################################################################
# CREATE AND FORMAT WORKSHEET #################################################
###############################################################################

    # PATH OF CONSUMER INPUT CSV ##############################################
    home_directory = os.path.expanduser("~")
    date_name = os.path.join(var_cons_prep.cons_dir, 'Regler_'+node.cons[x_node]+'_prepared.csv')
    print("Reading csv: "+date_name)

    # CREATE SHEET (EXCEL) ####################################################
    fileXLSX.create_sheet("C_"+str(node.cons[x_node]))
    fileXLSX.save(fileXLSX_name)
    sheet_cons = fileXLSX["C_"+str(node.cons[x_node])]

    # INSERT HEADLINE (EXCEL) #################################################
    sheet_cons_caption = ["Zeit", "V\u0307 [l/s]", "Q\u0307 [kW]", \
                          "t_VL [°C]", "t_RL [°C]", "p_VL [Pa]", "p_RL [Pa]", \
                          "t_last [s]", "CHECK"]
    for column_xlsx in range (1, len(sheet_cons_caption)+1):
        sheet_cons.cell(row = 1, column = column_xlsx).value = \
            sheet_cons_caption[column_xlsx-1]

    # FORMATTING (EXCEL) ######################################################
    bold_font = Font(bold = True)
    for cell in sheet_cons["1:1"]:
        cell.font = bold_font
    sheet_cons.freeze_panes = 'A2'
    sheet_cons.column_dimensions["A"].width = 20

###############################################################################
# READ MEASUREMENTS ###########################################################
###############################################################################

    # LOAD CSV FILE ###########################################################
    # Check for existence of csv file
    if os.path.isfile(date_name):

        # CSV FILE EXISTS #####################################################
        csv_exist = np.append(csv_exist, "v.")

        # INITIALIZATION ######################################################
        with open(date_name) as csvdatei:
            csv_reader_object = csv.reader(csvdatei, delimiter = ",")
            V_dot, temp_flow, temp_ret, Q_dot, p_flow, p_ret, t_last, \
                error = ([] for i in range(8))
            cntr = 0

            # CONSUMER MEASUREMENTS ###########################################
            for row in csv_reader_object:
                if csv_reader_object.line_num == 1:
                    continue
                if datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S") > \
                    var_sim.time_sim_end:
                        # NO MEASUREMENTS AVAILABLE IN SPECIFIED TIME RANGE: ##
                        # INITIATE SAME BEHAViOUR AS FOR NON-EXISTING CSV FILE #
                        if cntr == 0:
                            raise Exception("No measurement values available in the simulation period for CSV file " + str(date_name) + ". Please adjust the simulation period or set node " + str(node.nbr_orig_roman[x_node]) + " to inactive.")
                        break
                elif datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S") >= \
                    var_sim.time_sim_start:

                    # FILL ARRAYS #############################################
                    # Volumetric flow [l/h]
                    if row[4] != '':
                        V_dot = np.append(V_dot, \
                                              np.array([float(row[4])])/3600, \
                            axis = 0)
                    else:
                        V_dot = np.append(V_dot, np.array([0]), \
                                              axis = 0)

                    # Temperature of flow [°C]
                    if row[5] != '':
                        temp_flow = np.append(temp_flow, \
                            np.array([float(row[5])]), axis = 0)
                    else:
                        temp_flow = np.append(temp_flow, np.array([0]), \
                                              axis = 0)

                    # Temperature of return [°C]
                    if row[6] != '':
                        temp_ret = np.append(temp_ret, \
                            np.array([float(row[6])]), axis = 0)
                    else:
                        temp_ret = np.append(temp_ret, np.array([0]), axis = 0)

                    # Current heat flow/power [kW]
                    if row[1] != '':
                        Q_dot = np.append(Q_dot, \
                                          np.array([float(row[1])]), axis = 0)
                    else:
                        Q_dot = np.append(Q_dot, np.array([0]), axis = 0)

                    # Pressure of flow [kPa]
                    if row[7] != '':
                        p_flow = np.append(p_flow, \
                            np.array([float(row[7])])*1000, axis = 0)
                    else:
                        p_flow = np.append(p_flow, np.array([0]), axis = 0)

                    # Pressure of return [kPa]
                    if row[8] != '':
                        p_ret = np.append(p_ret, \
                            np.array([float(row[8])])*1000, axis = 0)
                    else:
                        p_ret = np.append(p_ret, np.array([0]), axis = 0)

                    # Time since last data transmission [s]
                    if row[9] != '':
                        t_last = np.append(t_last, \
                                           np.array([float(row[9])]), axis = 0)
                    else:
                        t_last = np.append(t_last, np.array([0]), axis = 0)

# BOOKMARK: Error recognition (consumers)
###############################################################################
# ERROR RECOGNITION ###########################################################
###############################################################################

                    # ERROR RECOGNITION THROUGH UNCHANGING MEASUREMENTS #######
                    # Set an error flag if the number of equal measurements for
                    # both flow and return temperature exceeds a certain value.
                    if var_gaps.error_detection == 0:
                        if cntr+1 >= var_gaps.nbr_equal_values_max:
                            if ([temp_flow[-1]]*var_gaps.nbr_equal_values_max \
                                == temp_flow[-var_gaps.nbr_equal_values_max::])\
                                .all() and ([temp_ret[-1]]*var_gaps.\
                                nbr_equal_values_max == temp_ret[-var_gaps.\
                                nbr_equal_values_max::]).all():
                                error = np.append(error, "FEHLER_m")
                            else:
                                error = np.append(error, "OKAY_m")
                        else:
                            error = np.append(error, "OKAY_m")

                    # TOP TRONIC ERROR RECOGNITION ############################
                    # Reads in results of external error recognition
                    elif var_gaps.error_detection == 1:
                        if float(row[9]) > 60*var_sim.delta_time_hyd:
                            error = np.append(error, "FEHLER_a")
                        else:
                            error = np.append(error, "OKAY_a")

                    # COMBINED ERROR RECOGNITION ##############################
                    # Reads in results of external error recognition, if they
                    # exist. Else, uses internal error recognition.
                    else:
                        try:
                            if float(row[9]) > 60*var_sim.delta_time_hyd:
                                error = np.append(error, "FEHLER_a")
                            else:
                                error = np.append(error, "OKAY_a")
                        except:
                            if cntr+1 >= var_gaps.nbr_equal_values_max:
                                if ([temp_flow[-1]]*var_gaps.\
                                    nbr_equal_values_max == \
                                    temp_flow[-var_gaps.\
                                    nbr_equal_values_max::]).all() and \
                                    ([temp_ret[-1]]*var_gaps.\
                                    nbr_equal_values_max == temp_ret[-var_gaps.\
                                    nbr_equal_values_max::]).all():
                                    error = np.append(error, "FEHLER_m")
                                else:
                                    error = np.append(error, "OKAY_m")
                            else:
                                error = np.append(error, "OKAY_m")

                    # RECOGNITION OF EMPTY ROWS ################################
                    # If the row is empty, raise an error flag "FEHLER_l"

                    if row[1] == '' and row[2] == '' and row[3] == '' and \
                        row[4] == '' and row[5] == '' and row[6] == '' and \
                        row[7] == '' and row[8] == '' and row[9] == '':
                            error[cntr] = "FEHLER_l"

###############################################################################
# WRITE MEASUREMENTS INTO EXCEL FILE ##########################################
###############################################################################

                    row_xlsx = cntr+2

                    # Time stamp [YYYY-MM-DD hh:mm:ss]
                    sheet_cons.cell(row = row_xlsx, column = 1).value = \
                        datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
                    sheet_cons.cell(row = row_xlsx, column = 1).style = \
                        var_misc.date_style

                    # Volumetric flow [l/s]
                    if row[4] != '':
                        sheet_cons.cell(row = row_xlsx, column = 2).value = \
                            float(row[4])/3600
                        sheet_cons.cell(row = row_xlsx, column = 2).\
                            number_format = "0.000"
                    else:
                        sheet_cons.cell(row = row_xlsx, column = 2).value = "-"

                    # Power [kW]
                    if row[1] != '':
                        sheet_cons.cell(row = row_xlsx, column = 3).value = \
                            float(row[1])
                    else:
                        sheet_cons.cell(row = row_xlsx, column = 3).value = "-"

                    # Flow temperature [°C]
                    if row[5] != '':
                        sheet_cons.cell(row = row_xlsx, column = 4).value = \
                            float(row[5])
                        sheet_cons.cell(row = row_xlsx, column = 4).\
                            number_format = "0.0"
                    else:
                        sheet_cons.cell(row = row_xlsx, column = 4).value = "-"

                    # Return temperature [°C]
                    if row[6] != '':
                        sheet_cons.cell(row = row_xlsx, column = 5).value = \
                            float(row[6])
                        sheet_cons.cell(row = row_xlsx, column = 5).\
                            number_format = "0.0"
                    else:
                        sheet_cons.cell(row = row_xlsx, column = 5).value = "-"

                    # Flow pressure [Pa]
                    if row[7] != '':
                        sheet_cons.cell(row = row_xlsx, column = 6).value = \
                            float(row[7])*1000
                        sheet_cons.cell(row = row_xlsx, column = 6).\
                            number_format = "# ##0"
                    else:
                        sheet_cons.cell(row = row_xlsx, column = 6).value = "-"

                    # RReturn pressure [Pa]
                    if row[8] != '':
                        sheet_cons.cell(row = row_xlsx, column = 7).value = \
                            float(row[8])*1000
                        sheet_cons.cell(row = row_xlsx, column = 7).\
                            number_format = "# ##0"
                    else:
                        sheet_cons.cell(row = row_xlsx, column = 7).value = "-"

                    # Time since last data transmission [s]
                    if row[9] != '':
                        sheet_cons.cell(row = row_xlsx, column = 8).value = \
                            float(row[9])
                        sheet_cons.cell(row = row_xlsx, column = 8).\
                            number_format = "# ##0"
                    else:
                        sheet_cons.cell(row = row_xlsx, column = 8).value = "-"

                    # Error flag
                    sheet_cons.cell(row = row_xlsx, column = 9).value = \
                        error[cntr]

                    # FORMATTING (EXCEL) ######################################
                    if error[cntr] == "FEHLER_m" or error[cntr] == "FEHLER_a" \
                        or error[cntr] == "FEHLER_l":
                        sheet_cons.cell(row = row_xlsx, column = 9).font = \
                            Font(color = "00FF0000")
                    else:
                        sheet_cons.cell(row = row_xlsx, column = 9).font = \
                            Font(color = "006400")

###############################################################################
# WRITE MEASUREMENTS INTO ARRAYS ##############################################
###############################################################################

                    # SET NEW INDICES #########################################
                    cntr += 1

        # FILL NODE OBJECT ####################################################
        node.V_dot.append(V_dot)
        node.temp_flow.append(temp_flow)
        node.temp_ret.append(temp_ret)
        node.Q_dot.append(Q_dot)
        node.p_flow.append(p_flow)
        node.p_ret.append(p_ret)
        node.error.append(error)

###############################################################################
# PROCEDURE FOR NON-EXISTING CSV-FILE #########################################
###############################################################################

    else:
        # CSV EXISTENCE INFO ##################################################
        csv_exist = np.append(csv_exist, "n.v.")

        # INITIALIZATION ####################################################
        time_cons = var_sim.time_sim_start
        V_dot, temp_flow, temp_ret, Q_dot, p_flow, p_ret, error = \
            ([] for i in range(7))
        row_xlsx = 2

        for cntr in range(0, len(var_sim.time_stamp)):

            # FILL WITH "-" (EXCEL) ###########################################
            sheet_cons.cell(row = cntr+2, column = 1).value = \
                var_sim.time_stamp[cntr]
            sheet_cons.cell(row = cntr+2, column = 1).style = \
                var_misc.date_style
            for column_xlsx in range (2, 8):
                sheet_cons.cell(row = cntr+2, column = column_xlsx).value = "-"

            # SETTING ERROR FLAGS AND FORMATTING (EXCEL) ######################
            sheet_cons.cell(row = cntr+2, column = 9).value = "FEHLER"
            sheet_cons.cell(row = cntr+2, column = 9).font = \
                Font(color = "00FF0000")

            # FILL ARRAYS ##############################################
            V_dot = np.append(V_dot, np.array([None]), axis = 0)
            temp_flow = np.append(temp_flow, np.array([None]), axis = 0)
            temp_ret = np.append(temp_ret, np.array([None]), axis = 0)
            Q_dot = np.append(Q_dot, np.array([None]), axis = 0)
            p_flow = np.append(p_flow, np.array([None]), axis = 0)
            p_ret = np.append(p_ret, np.array([None]), axis = 0)
            error = np.append(error, "FEHLER")

        # FILL NODE OBJECT ####################################################
        node.V_dot.append(V_dot)
        node.temp_flow.append(temp_flow)
        node.temp_ret.append(temp_ret)
        node.Q_dot.append(Q_dot)
        node.p_flow.append(p_flow)
        node.p_ret.append(p_ret)
        node.error.append(error)

###############################################################################
# OUTPUT ######################################################################
###############################################################################

    return (fileXLSX, node, csv_exist)

def read_cons_non_ID (fileXLSX, fileXLSX_name, x_node, var_sim, var_misc, node):
    """Fills the node object with blanks for the case of a consumer without ID.
    Called by read_data.

    :param fileXLSX: Active excel workbook
    :type fileXLSX: openpyxl.workbook.workbook obj.

    :param fileXLSX_name: Path of the active workbook
    :type fileXLSX_name: str

    :param x_node: Node number (w.r.t. coupling matrix) of the consumer
    :type x_node: int

    :param var_sim: Variables concerning the simulation
    :type var_sim: var_sim obj.

    :param var_misc: Miscellaneous variables
    :type var_misc: var_misc obj.

    :param node: Contains information about the nodes
    :type node: node obj.

    :return fileXLSX: Active excel workbook
    :rtype fileXLSX: openpyxl.workbook.workbook obj.

    :return node: Contains information about the nodes
    :rtype node: node obj.
    """    
    # CREATE WORKSHEET (EXCEL) ################################################
    fileXLSX.create_sheet("C_"+node.nbr_orig_roman[x_node])
    fileXLSX.save(fileXLSX_name)
    sheet_cons = fileXLSX["C_"+node.nbr_orig_roman[x_node]]

    # INSERT HEADER (EXCEL) ###################################################
    sheet_cons_caption = ["Zeit", "V\u0307 [l/s]", "Q\u0307 [kW]", \
                          "t_VL [°C]", "t_RL [°C]", "p_VL [Pa]", "p_RL [Pa]", \
                          "t_last [s]", "CHECK"]
    for column_xlsx in range (1, len(sheet_cons_caption)+1):
        sheet_cons.cell(row = 1, column = column_xlsx).value = \
            sheet_cons_caption[column_xlsx-1]

    # FORMATTING (EXCEL) ######################################################
    bold_font = Font(bold = True)
    for cell in sheet_cons["1:1"]:
        cell.font = bold_font
    sheet_cons.freeze_panes = 'A2'
    sheet_cons.column_dimensions["A"].width = 20

    # INITALIZATION ###########################################################
    V_dot, temp_flow, temp_ret, Q_dot, p_flow, p_ret, error \
        = ([] for i in range(7))

    for cntr in range(0, len(var_sim.time_stamp)):

        # FILLING WITH "-" (EXCEL) ############################################
        sheet_cons.cell(row = cntr+2, column = 1).value = \
            var_sim.time_stamp[cntr]
        sheet_cons.cell(row = cntr+2, column = 1).style = \
            var_misc.date_style
        for column_xlsx in range (2, 8):
            sheet_cons.cell(row = cntr+2, column = \
                column_xlsx).value = "-"

        # FORMATTING (EXCEL) ##################################################
        sheet_cons.cell(row = cntr+2, column = 9).value = \
            "FEHLER"
        sheet_cons.cell(row = cntr+2, column = 9).font = \
            Font(color = "00FF0000")

        # FILL ARRAYS #########################################################
        V_dot = np.append(V_dot, np.array([None]), axis = 0)
        temp_flow = np.append(temp_flow, np.array([None]), axis = 0)
        temp_ret = np.append(temp_ret, np.array([None]), axis = 0)
        Q_dot = np.append(Q_dot, np.array([None]), axis = 0)
        p_flow = np.append(p_flow, np.array([None]), axis = 0)
        p_ret = np.append(p_ret, np.array([None]), axis = 0)
        error = np.append(error, "FEHLER")

    # FILL LISTS ##############################################################
    node.V_dot.append(V_dot)
    node.temp_flow.append(temp_flow)
    node.temp_ret.append(temp_ret)
    node.Q_dot.append(Q_dot)
    node.p_flow.append(p_flow)
    node.p_ret.append(p_ret)
    node.error.append(error)

    # OUTPUT ##################################################################
    return (fileXLSX, node)

def read_feed (fileXLSX, fileXLSX_name, x_node, var_sim, var_misc, var_gaps, node):
    """Reads feeder data stored in separate csv files in "Daten" folder.
    Called by read_data.

    :param fileXLSX: Active excel workbook
    :type fileXLSX: openpyxl.workbook.workbook obj.

    :param fileXLSX_name: Path of the active workbook
    :type fileXLSX_name: str

    :param x_node: Node number (w.r.t. coupling matrix) of the feeder
    :type x_node: int

    :param var_sim: Variables concerning the simulation
    :type var_sim: var_sim obj.

    :param var_misc: Miscellaneous variables
    :type var_misc: var_misc obj.
    
    :param var_gaps: Variables concerning the gap filling
    :type var_gaps: var_gaps obj.

    :param node: Contains information about the nodes
    :type node: node obj.

    :return fileXLSX: Active excel workbook
    :rtype fileXLSX: openpyxl.workbook.workbook obj.

    :return node: Contains information about the nodes
    :rtype node: node obj.
    """    

    home_directory = os.path.expanduser("~")
    date_name = os.path.join(var_cons_prep.cons_dir, 'Router_WMZ_'+node.feed_in[x_node]+'_prepared.csv')
    print("Reading feeder csv: " + date_name)

    # CREATE SPREADSHEET (EXCEL) ##############################################
    fileXLSX.create_sheet("F_"+str(node.feed_in[x_node]))
    fileXLSX.save(fileXLSX_name)
    sheet_feed = fileXLSX["F_"+str(node.feed_in[x_node])]

    # LOAD CSV FILE ###########################################################
    with open(date_name) as csvdatei:
        csv_reader_object = csv.reader(csvdatei, delimiter = ",")

        # INSERT HEADER (EXCEL) ###############################################
        sheet_feed_caption = ["Zeit", "V\u0307 [l/s]", "Q\u0307 [kW]", \
                              "t_VL [°C]", "t_RL [°C]", "p_VL [Pa]", \
                              "p_RL [Pa]", "t_last [s]", "CHECK"]
        for x in range (0, len(sheet_feed_caption)):
            sheet_feed.cell(row = 1, column = x+1).value = \
                sheet_feed_caption[x]

        # FORMATTING (EXCEL) ##################################################
        bold_font = Font(bold = True)
        for cell in sheet_feed["1:1"]:
            cell.font = bold_font
        sheet_feed.freeze_panes = 'A2'
        sheet_feed.column_dimensions["A"].width = 20

        # MEASUREMENTS OF FEEDER ##############################################
        Q_dot, V_dot, temp_flow, temp_ret, p_flow, p_ret, t_last, \
            error = ([] for i in range(8))
        cntr = 0
        for row in csv_reader_object:
            # Skip first row
            if csv_reader_object.line_num == 1:
                continue
            if datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S') > \
                var_sim.time_sim_end:
                break
            elif datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S') >= \
                var_sim.time_sim_start:

                # FILL ARRAYS #################################################
                # Current power [kW]
                if row[1] != '':
                    Q_dot = np.append(Q_dot, \
                        np.array([float(row[1])]), axis = 0)
                else:
                    Q_dot = np.append(Q_dot, np.array([0]), axis = 0)
                
                # Volumetric flow [l/h]
                if row[4] != '':
                    V_dot = np.append(V_dot, \
                        np.array([float(row[4])/3600]), axis = 0)
                else:
                    V_dot = np.append(V_dot, np.array([0]), axis = 0)
                
                # Temperature of flow [°C]
                if row[5] != '':
                    temp_flow = np.append(temp_flow, \
                        np.array([float(row[5])]), axis = 0)
                else:
                    temp_flow = np.append(temp_flow, np.array([0]), \
                        axis = 0)
                
                # Temperature of return [°C]
                if row[6] != '':
                    temp_ret = np.append(temp_ret, \
                        np.array([float(row[6])]), axis = 0)
                else:
                    temp_ret = np.append(temp_ret, \
                        np.array([0]), axis = 0)
                
                # Pressure of flow [kPa]
                if row[7] != '':
                    p_flow = np.append(p_flow, \
                        np.array([float(row[7])*1000]), axis = 0)
                else:
                    p_flow = np.append(p_flow, \
                        np.array([0]), axis = 0)
                
                # Pressure of return [kPa]
                if row[8] != '':
                    p_ret = np.append(p_ret, \
                        np.array([float(row[8])*1000]), axis = 0)
                else:
                    p_ret = np.append(p_ret, \
                        np.array([0]), axis = 0)
                    
                # Time since last data transmission [s]
                if row[9] != '':
                    t_last = np.append(t_last, \
                        np.array([float(row[9])]), axis = 0)
                else:
                    t_last = np.append(t_last, \
                        np.array([0]), axis = 0)

# BOOKMARK: Error recognition (feeders)
                # ERROR DETECTION #############################################
                # Error recognition by unchanging measurements
                if var_gaps.error_detection == 0:
                    if cntr+1 >= var_gaps.nbr_equal_values_max:
                        if ([temp_flow[-1]]*\
                            var_gaps.nbr_equal_values_max == \
                            temp_flow[-var_gaps.\
                            nbr_equal_values_max::]).all() and \
                            ([temp_ret[-1]]*var_gaps.\
                            nbr_equal_values_max == \
                            temp_ret[-var_gaps.\
                            nbr_equal_values_max::]).all():
                            error = np.append(error, "FEHLER_m")
                        else:
                            error = np.append(error, "OKAY_m")
                    else:
                        error = np.append(error, "OKAY_m")

                # Top tronic error recognition
                elif var_gaps.error_detection == 1:
                    if float(row[9]) > 60*var_sim.delta_time_hyd:
                        error = np.append(error, "FEHLER_a")
                    else:
                        error = np.append(error, "OKAY_a")

                # Combined error recognition
                else:
                    try:
                        if float(row[9]) > 60*var_sim.delta_time_hyd:
                            error = np.append(error, "FEHLER_a")
                        else:
                            error = np.append(error, "OKAY_a")
                    except:
                        if cntr+1 >= var_gaps.nbr_equal_values_max:
                            if ([temp_flow[-1]]*\
                                var_gaps.nbr_equal_values_max == \
                                temp_flow[-var_gaps.\
                                nbr_equal_values_max::]).all() and \
                                ([temp_ret[-1]]*var_gaps.\
                                nbr_equal_values_max == \
                                temp_ret[-var_gaps.\
                                nbr_equal_values_max::]).all():
                                error = np.append(error, "FEHLER_m")
                            else:
                                error = np.append(error, "OKAY_m")
                        else:
                            error = np.append(error, "OKAY_m")

                # RECOGNITION OF EMPTY ROWS ################################
                # If the row is empty, raise an error flag "FEHLER_l"

                if row[1] == '' and row[2] == '' and row[3] == '' and \
                   row[4] == '' and row[5] == '' and row[6] == '' and \
                   row[7] == '' and row[8] == '' and row[9] == '':
                    error[cntr] = "FEHLER_l"


                # FILL EXCEL SHEET ############################################
                row_xlsx = cntr+2

                # Time stamp [YYYY-MM-DD hh:mm:ss]
                sheet_feed.cell(row = row_xlsx, column = 1).value = \
                    datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
                sheet_feed.cell(row = row_xlsx, column = 1).style = \
                    var_misc.date_style

                # Volumetric flow [l/s]
                if row[4] != '':
                    sheet_feed.cell(row = row_xlsx, \
                        column = 2).value = float(row[4])/3600
                    sheet_feed.cell(row = row_xlsx, column = 2).\
                        number_format = "0.000"
                else:
                    sheet_feed.cell(row = row_xlsx, \
                        column = 2).value = "-"

                # Power [kW]
                if row[1] != '':
                    sheet_feed.cell(row = row_xlsx, \
                        column = 3).value = float(row[1])
                    sheet_feed.cell(row = row_xlsx, column = 3).\
                        number_format = "# ##0"
                else:
                    sheet_feed.cell(row = row_xlsx, \
                        column = 3).value = "-"

                # Flow temperature [°C]
                if row[5] != '':
                    sheet_feed.cell(row = row_xlsx, \
                        column = 4).value = float(row[5])
                    sheet_feed.cell(row = row_xlsx, column = 4).\
                        number_format = "0.0"
                else:
                    sheet_feed.cell(row = row_xlsx, \
                        column = 4).value = "-"

                # Return temperature [°C]
                if row[6] != '':
                    sheet_feed.cell(row = row_xlsx, \
                        column = 5).value = float(row[6])
                    sheet_feed.cell(row = row_xlsx, column = 5).\
                        number_format = "0.0"
                else:
                    sheet_feed.cell(row = row_xlsx, \
                        column = 5).value = "-"

                # Flow pressure [Pa]
                if row[7] != '':
                    sheet_feed.cell(row = row_xlsx, \
                        column = 6).value = float(row[7])*1000
                    sheet_feed.cell(row = row_xlsx, column = 6).\
                        number_format = "# ##0"
                else:
                    sheet_feed.cell(row = row_xlsx, \
                        column = 6).value = "-"

                # Return pressure [Pa]
                if row[8] != '':
                    sheet_feed.cell(row = row_xlsx, \
                        column = 7).value = float(row[8])*1000
                    sheet_feed.cell(row = row_xlsx, column = 7).\
                        number_format = "# ##0"
                else:
                    sheet_feed.cell(row = row_xlsx, \
                        column = 7).value = "-"

                # Time since last data transmission [s]
                if row[9] != '':
                    sheet_feed.cell(row = row_xlsx, \
                        column = 8).value = float(row[9])
                    sheet_feed.cell(row = row_xlsx, column = 8).\
                        number_format = "# ##0"
                else:
                    sheet_feed.cell(row = row_xlsx, \
                        column = 8).value = "-"
                sheet_feed.cell(row = cntr+2, column = 9).value = \
                    error[cntr]

                # FORMATTING (EXCEL) ##########################################
                if error[cntr] == "FEHLER_m" or error[cntr] == \
                    "FEHLER_a" or error[cntr] == "FEHLER_l":
                    sheet_feed.cell(row = row_xlsx, column = 9).font \
                        = Font(color = "00FF0000")
                else:
                    sheet_feed.cell(row = row_xlsx, column = 9).font \
                        = Font(color = "006400")

                # SET NEW INDICES #############################################
                cntr += 1

        # FILL LISTS ##########################################################
        node.Q_dot_feed.append(Q_dot)
        node.V_dot_feed.append(V_dot)
        node.temp_flow_feed.append(temp_flow)
        node.temp_ret_feed.append(temp_ret)
        node.p_flow_feed.append(p_flow)
        node.p_ret_feed.append(p_ret)
        node.error_feed.append(error)
    
    return (fileXLSX, node)