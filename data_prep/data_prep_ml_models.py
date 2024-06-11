# -*- coding: utf-8 -*-
"""
Author: 4wardEnergy Research GmbH
Date: 2024-05-14
Version: 1.0

This subroutine prepares the data for machine learning (ML) models and 
automatically selects the gap-filling mode based on data availability. 
It includes functions for data interpolation, autocorrelation analysis, 
and ML model training and evaluation using AutoML. The results, including 
performance metrics and plots, are saved to the topology file for further 
analysis and decision-making.
"""
"""
Copyright (C) 2024  4wardEnergy Research GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from supervised.automl import AutoML
import matplotlib.pyplot as plt
from sklearn.metrics import mean_absolute_error
from matplotlib import dates as mdates
from sklearn.metrics import mean_squared_error
from matplotlib import gridspec
from winsound import Beep
import termcolor
import shutil
import json
from scipy.signal import find_peaks
import matplotlib
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule
from sklearn.metrics import r2_score
import datetime
import time
import importlib

# Reload options to make sure that the latest changes are applied.
import options as options
importlib.reload(options)
from options import *

matplotlib.use('Agg')

# BOOKMARK: Helper functions ##################################################
def print_green(text):
    print(termcolor.colored(text, "green"))

def print_red(text):
    print(termcolor.colored(text, "red"))

def interpolate_pandas_time_series(df, timestamps, ffill=True):
    """Interpolate missing values in a pandas DataFrame based on a given set of timestamps.

    :param df: The DataFrame containing the time series data.
    :type df: pandas.DataFrame
    :param timestamps: The list of timestamps to be used for interpolation.
    :type timestamps: list
    :param ffill: If True, forward fill missing values. If False, interpolate using linear method.
    :type ffill: bool, optional

    :returns: The DataFrame with interpolated values.
    :rtype: pandas.DataFrame
    """

    # Step 1: Merge with Existing Data
    # Create a DataFrame from timestamps
    complete_data = pd.DataFrame(timestamps, columns=['Zeitstempel'])
    # Merge with y_test_wb
    df_int = pd.merge(complete_data, df, on='Zeitstempel', how='left')
    # Step 2: Interpolate Missing Values
    if ffill:
        df_int.ffill(inplace=True)
    else:
        df_int.interpolate(method="linear",inplace=True)
    return df_int

def autocorrelation(time_series, max_lag=250):
    """Calculate the autocorrelation of a time series up to a specified lag.

    :param time_series: Input time series as a numpy array.
    :param max_lag: Maximum lag (number of steps) to compute the autocorrelation for.
    :return: Autocorrelation function of the time series up to max_lag.
    """
    n = len(time_series)
    time_series = time_series - np.mean(time_series)
    result = np.correlate(time_series, time_series, mode='full')[-n:]
    acf = result[:max_lag+1] / result[0]
    return acf

def detect_peaks(acf, prominence):
    """Detect peaks in the autocorrelation function.

    :param acf: Autocorrelation function as a numpy array.
    :param prominence: Minimum peak prominence.
    :return: Indices of peaks that are above the threshold.
    """
    peaks, _ = find_peaks(acf, prominence=prominence, distance=80)
    return peaks

def plot_autocorrelation(acf, peaks):
    """Plot the autocorrelation function with detected peaks and threshold line.

    :param acf: Autocorrelation function as a numpy array.
    :param peaks: Indices of detected peaks.
    """
    plt.figure(figsize=(10, 6))
    lags = np.arange(len(acf))
    plt.plot(lags, acf, label='Autokorrelation')
    plt.plot(peaks, acf[peaks], "x", label='Erkannte Peaks')
    plt.xlabel('Lag')
    plt.ylabel('Autokorrelation')
    plt.title('Autokorrelationsfunktion')
    plt.legend()
    plt.savefig("Autocorrelation.png", dpi=300, bbox_inches='tight')
    plt.close()


def ts_autocorr(time_series):
    """Perform autocorrelation analysis on a time series and check for specific lags.

    :param time_series: Input time series as a numpy array.
    :type time_series: numpy.ndarray
    :return: True if peaks are detected at specific lags, False otherwise.
    """
    # Calculate autocorrelation up to a specified lag
    acf = autocorrelation(time_series, 220)

    # Calculate min and max of acf
    acf_min = np.min(acf)
    acf_max = np.max(acf)
    acf_range = acf_max - acf_min

    # Smooth acf
    acf = pd.Series(acf).rolling(window=5).mean().values
    # Detect peaks
    peaks = detect_peaks(acf, prominence=acf_range*0.1)

    # Plot autocorrelation with threshold
    plot_autocorrelation(acf, peaks)

    # Check for specific lags (94...98 and 190...194) in the date
    peak_at_96 = (94 in peaks) or (95 in peaks) or (96 in peaks) \
                or (97 in peaks) or (98 in peaks)
    peak_at_192 = (190 in peaks) or (191 in peaks) or (192 in peaks) \
                or (193 in peaks) or (194 in peaks)

    return peak_at_96 and peak_at_192

def interpolate_and_predict(X_test, y_test, date_cut, automl):
    """Interpolate missing values in the test set and makes predictions using the AutoML model.

    :param X_test: Test set features.
    :type X_test: pandas.DataFrame
    :param y_test: Test set labels.
    :type y_test: pandas.DataFrame
    :param date_cut: Date of the last timestamp in the training set.
    :type date_cut: pandas.Timestamp
    :param automl: AutoML model.
    :type automl: supervised.automl.AutoML

    :return: Predictions, interpolated predictions, ground truth, interpolated ground truth, 
             test set features, interpolated test set features, time delta.
    :rtype: tuple
    """
    # INTERPOLATION ###########################################################
    # Interpolate missing values to get to a whole day of data
    # Create Timestamps
    # Find first and last timestamp in series
    first_timestamp = y_test["Zeitstempel"].iloc[0]
    last_timestamp = y_test["Zeitstempel"].iloc[-1]
    timestamps = pd.date_range(start=first_timestamp, end=last_timestamp, freq="15min")

    X_test_int = interpolate_pandas_time_series(X_test, timestamps, ffill=True)
    y_test_int = interpolate_pandas_time_series(y_test, timestamps, ffill=True)

    # Constrain X_test and y_test to 50 days
    delta = pd.Timedelta("50 days")
    one_week = pd.Timedelta("7 days")
    X_test_cut_int = X_test_int[(X_test_int["Zeitstempel"] >= date_cut + one_week) & (X_test_int["Zeitstempel"] <= date_cut + one_week + delta)]
    y_test_cut_int = y_test_int[(y_test_int["Zeitstempel"] >= date_cut + one_week) & (y_test_int["Zeitstempel"] <= date_cut + one_week + delta)]
    X_test_cut = X_test[(X_test["Zeitstempel"] >= date_cut + one_week) & (X_test["Zeitstempel"] <= date_cut + one_week + delta)]
    y_test_cut = y_test[(y_test["Zeitstempel"] >= date_cut + one_week) & (y_test["Zeitstempel"] <= date_cut + one_week + delta)]

    # Take data from the week before
    y_test_wb = y_test[(X_test["Zeitstempel"] >= date_cut) & (X_test["Zeitstempel"] <= date_cut + delta)]
    # Interpolate it linearly
    timestamps_wb = pd.date_range(start=date_cut, end=date_cut + delta, freq="15min")
    y_test_wb_int = interpolate_pandas_time_series(y_test_wb, timestamps_wb, ffill=False)
    # Set timestamps to 1 week later
    y_test_wb_int["Zeitstempel"] = y_test_wb_int["Zeitstempel"] + pd.Timedelta("7 days")

    predictions_int = automl.predict(X_test_cut_int)
    # Create a dataframe with the predictions and the timestamps
    predictions_int = pd.DataFrame(predictions_int, columns=["predictions"])
    predictions_int["Zeitstempel"] = X_test_cut_int["Zeitstempel"]
    # Merge with X_test to only retain samples with ground truth
    predictions = pd.merge(predictions_int, X_test_cut, on="Zeitstempel", how="right")

    return predictions, predictions_int, y_test_cut, y_test_wb_int, X_test_cut, X_test_cut_int, delta

def format_workbook(excel_name):
    """Format the Excel workbook containing the results of the ML analysis.

    :param excel_name: Name of the Excel file.
    :type excel_name: str
    """

    workbook = openpyxl.load_workbook(excel_name)
    worksheet = workbook.active
    
    for cell in worksheet['O']:
        # SKip first row
        if cell.row != 1:
            # Skip formatting if cell already has a hyperlink
            if "HYPERLINK" in str(cell.value):
                continue
            if cell.value is None:
                continue
            linkpath = str(cell.value)
            cell.value = '=HYPERLINK("{}", "{}")'.format("file:///"+linkpath, "Plot")


    for cell in worksheet['P']:
        # SKip first row
        if cell.row != 1:
            if "HYPERLINK" in str(cell.value):
                continue
            if cell.value is None:
                continue
            linkpath = str(cell.value)
            cell.value = '=HYPERLINK("{}", "{}")'.format("file:///"+linkpath, "Plot")
    
    for cell in worksheet['Q']:
        # SKip first row
        if cell.row != 1:
            if "HYPERLINK" in str(cell.value):
                continue
            if cell.value is None:
                continue
            linkpath = str(cell.value)
            cell.value = '=HYPERLINK("{}", "{}")'.format("file:///"+linkpath, "Plot")

    # Format Links
    for cell in worksheet['O']:
        if cell.row != 1:
            cell.font = openpyxl.styles.Font(underline="single", color="0000FF")
    for cell in worksheet['P']:
        if cell.row != 1:
            cell.font = openpyxl.styles.Font(underline="single", color="0000FF")
    for cell in worksheet['Q']:
        if cell.row != 1:
            cell.font = openpyxl.styles.Font(underline="single", color="0000FF")
    
    # Set column widths
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        worksheet.column_dimensions[col].width = 15
    for col in ["N", "O", "P", "Q"]:
        worksheet.column_dimensions[col].width = 10
    
    # Round to 2 decimals
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        for cell in worksheet[col][1:]:
            cell.number_format = "0.00"

    # Conditional formatting for NMAE difference
    for col in ["L"]:
        for cell in worksheet[col][1:]:
            # Define the range for conditional formatting for the entire column (excluding the header)
            column_range = f"{col}2:{col}{worksheet.max_row}"

            # Create a 3-color scale rule: green for negative, white for -0.05, red for positive
            rule1 = ColorScaleRule(start_type='min', start_color='002fad53',
                                mid_type='num', mid_value=-0.05, mid_color='FFFFFFFF',
                                end_type='max', end_color='00ad3e2f')

            # Add the rule to the worksheet
            worksheet.conditional_formatting.add(column_range, rule1)

    # Sort by NMAE difference
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.auto_filter.add_sort_condition("L2:L" + str(worksheet.max_row))

    # Save workbook
    workbook.save(excel_name)   


# BOOKMARK: ML main functions #################################################

def data_preparation(cons_id):
    """Main routine.
    Prepare the data for machine learning (ML) models and automatically 
    select the gap-filling mode based on data availability.

    :param cons_id: Consumer ID.
    :type cons_id: str
    """
    # DATA READ-IN ############################################################
    # Read data
    # Choose cons. ID
    file_name = "Regler_" + cons_id + "_prepared.csv"
    file_path = os.path.join(var_cons_prep.cons_dir, file_name)

    df = pd.read_csv(
        file_path,
        skipinitialspace=True,
    )

    # Check max. power from excel
    excel = pd.read_excel(var_cons_list_analysis.topology_file, sheet_name="Knoten", dtype={"Abnehmer": str})
    max_power = excel[excel["Abnehmer"] == cons_id]["Nennleistung [kW]"].values[0]

    # Drop rows where "akt. Leistung(kW)" is higher than max. power
    df = df[df["akt. Leistung(kW)"] <= max_power*1.25]

    # Parse to datetime
    df["Zeitstempel"] = pd.to_datetime(df["Zeitstempel"])

    # Drop rows with unchanging values
    equal_values_max_timesteps = int(var_gaps.equal_values_max_min / 15)

    # Flag rows where "Vorlauftemperatur (°C)" has not changed
    df['Flag Unverändert'] = \
        (df['Vorlauftemperatur (°C)'] == df['Vorlauftemperatur (°C)'].shift(periods=equal_values_max_timesteps - 1)) & \
        (df['Ruecklauftemperatur (°C)'] == df['Ruecklauftemperatur (°C)'].shift(periods=equal_values_max_timesteps - 1))

    # Drop rows where "Vorlauftemperatur (°C)" has not changed
    df = df[~df['Flag Unverändert']]
    df = df.drop(columns=['Flag Unverändert'])

    # Drop unnecessary columns
    df = df[df.columns[[0,1,10]]]
    # Make columns for hour, day, week, month
    # df["hour"] = df["Zeitstempel"].dt.hour
    # df["day"] = df["Zeitstempel"].dt.weekday
    # df["week"] = df["Zeitstempel"].dt.isocalendar().week
    # df ["month"] = df["Zeitstempel"].dt.month

    # Fill NaN values in temperature with temperature data from the weather file
    # Read weather file
    weather = pd.read_csv(var_load_profiles.weather_file, skipinitialspace=True)
    # Rename time column to "Zeitstempel"
    weather = weather.rename(columns={"time": "Zeitstempel"})
    # Drop "station" column
    weather = weather.drop(columns=["station"])
    # Remove last 6 characters from "Zeitstempel" to match the format of the df
    weather["Zeitstempel"] = weather["Zeitstempel"].str[:-6]
    # Parse to datetime
    weather["Zeitstempel"] = pd.to_datetime(weather["Zeitstempel"])
    
    # Interpolate weather data to match the timestamps of df
    weather = weather.set_index("Zeitstempel")
    weather = weather.resample("15T").interpolate(method="linear")
    weather.index = weather.index + pd.Timedelta("5 minutes")
    weather = weather.reset_index()

    # Merge with df
    df = pd.merge(df, weather, on="Zeitstempel", how="left")
    # Fill values where "Aussentemperatur (°C)" is NaN with "TTX" from the weather file
    df["Aussentemperatur (°C)"] = df["Aussentemperatur (°C)"].fillna(df["TTX"])

    # Delete "TTX" column
    df = df.drop(columns=["TTX"])


    # Drop rows with NaN values
    df = df.dropna()

    # Check if there's anything left
    if df.empty:
        raise ValueError(f"Abnehmer {cons_id}: Keine Daten vorhanden. Überspringe Abnehmer.")

    # Get date of most recent timestamp
    date_max = df["Zeitstempel"].max()
    # Set date_max to this day, 00:00:00
    date_max = pd.to_datetime(str(date_max.date()) + " 00:00:00")
    
    # Set date_cut to the last 2 months of winter
    # If the month of date_max is 12,1,2,3,4 or 5
    if date_max.month in [12,1,2,3,4,5]:
        # Set date_cut to 2 months before date_max
        date_cut = date_max - pd.DateOffset(months=2)
    else:
        # Set date_cut to first of march last year
        date_cut = pd.to_datetime(str(date_max.year - 1) + "-03-01 00:00:00")

    # Set time of date_cut to 00:05
    date_cut = date_cut + pd.Timedelta("5 minutes")

    #Training set: Up to date_cut, Test set: From date_cut

    df_train = df[df["Zeitstempel"] < date_cut]
    X_train = df_train[df_train.columns[[0,2]]]
    y_train = df_train[["Zeitstempel","akt. Leistung(kW)"]]

    df_test = df[df["Zeitstempel"] >= date_cut]
    X_test = df_test[df_test.columns[[0,2]]]
    y_test = df_test[["Zeitstempel","akt. Leistung(kW)"]]

    return X_train, y_train, X_test, y_test, df, date_cut

# BOOKMARK: Function to create ML models
def preliminary_analysis(cons_id):
    """Perform preliminary analysis on the data and create ML models.

    :param cons_id: Consumer ID.
    :type cons_id: str
    """

    # DATA READ-IN ############################################################
    try:
        X_train, y_train, X_test, y_test, df, date_cut = data_preparation(cons_id)
    except:
        print_red(f"Abnehmer {cons_id}: Einlesen fehlgeschlagen. Überspringe Abnehmer.")
        return False


    # PLOT TRAINING AND TEST SET ##############################################
    fig = plt.figure(figsize = (16,9))
    plt.suptitle("Vorhandene Daten für Abnehmer " + cons_id)
    # set height ratios for subplots
    gs = gridspec.GridSpec(2, 1, height_ratios=[2, 1]) 

    # the first subplot
    ax0 = plt.subplot(gs[0])
    line0, = ax0.plot(y_train["Zeitstempel"], y_train["akt. Leistung(kW)"], label="Trainingsdatensatz")
    line1, = ax0.plot(y_test["Zeitstempel"], y_test["akt. Leistung(kW)"], label="Testdatensatz")
    # y-axis label
    ax0.set_ylabel("Leistung (kW)")
    # Set up grid: every month
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator(interval=1))
    plt.grid()

    # the second subplot
    # shared axis X
    ax1 = plt.subplot(gs[1], sharex = ax0)
    line2, = ax1.plot(df["Zeitstempel"], df["Aussentemperatur (°C)"], label="Außentemperatur", color="C3")
    plt.setp(ax0.get_xticklabels(), visible=False)
    # Rotate x-axis labels
    plt.xticks(rotation=45, ha="right")
    # y-axis label
    ax1.set_ylabel("Außentemperatur (°C)")
    # Set up grid: every month
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator(interval=1))
    plt.grid()

    # put legends
    ax0.legend()

    # remove vertical gap between subplots
    plt.subplots_adjust(hspace=.0)
    plt.savefig("training_test_set.png", dpi=300, bbox_inches='tight')
    plt.close()

    # AUTOCORRELATION PLOTS ###################################################
    # Plot autocorrelation
    try:
        autocorr_peaks_exist = ts_autocorr(y_train["akt. Leistung(kW)"].values)
    except:
        print_red(f"Abnehmer {cons_id}: Autokorrelation konnte nicht berechnet werden.")
        return False
    
    return autocorr_peaks_exist
    
def create_ml_models(cons_id):
    """Create machine learning (ML) models for the given consumer ID.

    :param cons_id: Consumer ID.
    :type cons_id: str
    """
    # READ DATA ################################################################
    X_train, y_train, X_test, y_test, df, date_cut = data_preparation(cons_id)

    # AUTOML FITTING ##########################################################
    # BOOKMARK: AUTOML FITTING
    automl = AutoML(mode="Perform", explain_level=1, eval_metric="mae", algorithms=["Baseline", "LightGBM", "Xgboost", "CatBoost", "Neural Network"], verbose=0) # AutoML(mode="Explain", explain_level=1, eval_metric="mae", algorithms = ["Decision Tree"]) #
    automl.fit(X_train, y_train["akt. Leistung(kW)"])
    
    # INTERPOLATION AND PREDICTION ############################################
    predictions, predictions_int, y_test_cut, y_test_wb_int, X_test_cut, X_test_cut_int, delta = \
        interpolate_and_predict(X_test, y_test, date_cut, automl)

    # BASELINE MODEL EVAL. ####################################################
    # Merge y_test_wb_int with y_test_cut so that only real values are retained
    y_test_wb_int_eval = pd.merge(y_test_wb_int, y_test_cut, on="Zeitstempel", how="right")

    # Compute MAE
    #Drop NaN values
    y_test_wb_int_eval = y_test_wb_int_eval.dropna()
    mae_bl = mean_absolute_error(y_test_wb_int_eval["akt. Leistung(kW)_x"].values, y_test_wb_int_eval["akt. Leistung(kW)_y"].values)

    # Compute RMSE
    rmse_bl = mean_squared_error(y_test_wb_int_eval["akt. Leistung(kW)_x"].values, y_test_wb_int_eval["akt. Leistung(kW)_y"].values, squared=False)

    # AUTOML MODEL
    # Compute MAE
    mae_automl = mean_absolute_error(predictions["predictions"].values, y_test_cut["akt. Leistung(kW)"].values)

    # Compute RMSE
    rmse_automl = mean_squared_error(predictions["predictions"].values, y_test_cut["akt. Leistung(kW)"].values, squared=False)
    
    # OVERVIEW PLOTS ##########################################################
    # Plot predictions
    # Two subplots
    plt.figure(figsize=(18,13))
    plt.subplot(2,1,1)
    plt.plot(X_test_cut["Zeitstempel"], y_test_cut["akt. Leistung(kW)"].values, label="Messwerte",marker=".")
    plt.plot(X_test_cut_int["Zeitstempel"],predictions_int["predictions"], label="Prognose", color = "C1")

    plt.title(f"AutoML-Prognosen für {cons_id}: RMSE = {round(rmse_automl,2)} kW | MAE = {round(mae_automl,2)} kW")
    plt.xlabel("Zeit")
    plt.ylabel("Leistung (kW)")
    #Format x-axis labels
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%A %d.%m.'))
    plt.gca().xaxis.set_minor_formatter(mdates.DateFormatter('%A %d.%m.'))
    plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=7))
    plt.gca().xaxis.set_minor_locator(mdates.DayLocator(interval=1))
    plt.xticks(rotation=45, ha="right")
    # Rotate minor x-axis labels
    for tick in plt.gca().xaxis.get_minor_ticks():
        tick.label1.set_horizontalalignment('right')
        tick.label1.set_rotation(45)
    # Set tight xlim
    plt.xlim(date_cut + pd.Timedelta("7 days"), date_cut + pd.Timedelta("7 days") + delta)
    plt.grid()
    plt.legend()

    # Plot predictions
    plt.subplot(2,1,2)
    plt.plot(X_test_cut["Zeitstempel"], y_test_cut["akt. Leistung(kW)"].values, label="Messwerte",marker=".")
    plt.plot(y_test_wb_int["Zeitstempel"],y_test_wb_int["akt. Leistung(kW)"].values, label="Vorwoche", color="C1")

    plt.title(f"Baseline für {cons_id}: RMSE = {round(rmse_bl,2)} kW | MAE = {round(mae_bl,2)} kW")
    plt.xlabel("Zeit")
    plt.ylabel("Leistung (kW)")
    #Format x-axis labels
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%A %d.%m.'))
    plt.gca().xaxis.set_minor_formatter(mdates.DateFormatter('%A %d.%m.'))
    plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=7))
    plt.gca().xaxis.set_minor_locator(mdates.DayLocator(interval=1))
    plt.xticks(rotation=45, ha="right")
    # Rotate minor x-axis labels
    for tick in plt.gca().xaxis.get_minor_ticks():
        tick.label1.set_horizontalalignment('right')
        tick.label1.set_rotation(45)
    # Set tight xlim
    plt.xlim(date_cut, date_cut + delta)
    plt.grid()
    plt.legend()
    plt.tight_layout()
    plt.savefig("Best model predictions overview.png", dpi=300, bbox_inches='tight')
    plt.close()

    # DETAIL PLOTS
    start_plot = date_cut + pd.Timedelta("14 days")
    delta_plot = pd.Timedelta("7 days")
    end_plot = start_plot + delta_plot

    # Plot predictions
    # Two subplots
    plt.figure(figsize=(18,13))
    plt.subplot(2,1,1)
    plt.plot(X_test_cut["Zeitstempel"], y_test_cut["akt. Leistung(kW)"].values, label="Messwerte",marker=".")
    plt.plot(X_test_cut_int["Zeitstempel"],predictions_int["predictions"], label="Prognose", color = "C1")
    # Rotate x-axis labels
    plt.xticks(rotation=45, ha="right")
    plt.title(f"AutoML-Prognosen für {cons_id}: RMSE = {round(rmse_automl,2)} kW | MAE = {round(mae_automl,2)} kW")
    plt.xlabel("Zeit")
    plt.ylabel("Leistung (kW)")
    #Format x-axis labels
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%A %d.%m.'))
    plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=1))
    plt.gca().xaxis.set_minor_locator(mdates.HourLocator(interval=1))
    plt.grid()
    plt.legend()
    # Set xlim
    plt.xlim(start_plot, end_plot)

    # Plot predictions
    plt.subplot(2,1,2)
    plt.plot(X_test_cut["Zeitstempel"], y_test_cut["akt. Leistung(kW)"].values, label="Messwerte",marker=".")
    plt.plot(y_test_wb_int["Zeitstempel"],y_test_wb_int["akt. Leistung(kW)"].values, label="Vorwoche", color="C1")
    # Rotate x-axis labels
    plt.xticks(rotation=45, ha="right")
    plt.title(f"Baseline für {cons_id}: RMSE = {round(rmse_bl,2)} kW | MAE = {round(mae_bl,2)} kW")
    plt.xlabel("Zeit")
    plt.ylabel("Leistung (kW)")
    #Format x-axis labels
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%A %d.%m.'))
    plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=1))
    plt.gca().xaxis.set_minor_locator(mdates.HourLocator(interval=1))
    plt.grid()
    plt.legend()
    # Set xlim
    plt.xlim(start_plot, end_plot)
    plt.tight_layout()
    plt.savefig("Best model predictions detail.png", dpi=300, bbox_inches='tight')
    plt.close()

    # BASELINE MODEL EVAL. ####################################################
    # Merge y_test_wb_int with y_test_cut so that only real values are retained
    y_test_wb_int_eval = pd.merge(y_test_wb_int, y_test_cut, on="Zeitstempel", how="right")

    # Compute MAE
    #Drop NaN values
    y_test_wb_int_eval = y_test_wb_int_eval.dropna()
    mae_bl = mean_absolute_error(y_test_wb_int_eval["akt. Leistung(kW)_x"].values, y_test_wb_int_eval["akt. Leistung(kW)_y"].values)

    # Compute RMSE
    rmse_bl = mean_squared_error(y_test_wb_int_eval["akt. Leistung(kW)_x"].values, y_test_wb_int_eval["akt. Leistung(kW)_y"].values, squared=False)

    # Compute R2
    r2_bl = r2_score(y_test_wb_int_eval["akt. Leistung(kW)_x"].values, y_test_wb_int_eval["akt. Leistung(kW)_y"].values)

    # Compute MAPE
    nmae_bl = mae_bl/y_test_wb_int_eval["akt. Leistung(kW)_x"].values.mean()

    # AUTOML MODEL
    # Compute MAE
    mae_automl = mean_absolute_error(y_test_cut["akt. Leistung(kW)"].values, predictions["predictions"].values)

    # Compute RMSE
    rmse_automl = mean_squared_error(y_test_cut["akt. Leistung(kW)"].values, predictions["predictions"].values, squared=False)

    # Compute R2
    r2_automl = r2_score(y_test_cut["akt. Leistung(kW)"].values, predictions["predictions"].values)

    # Compute MAPE
    nmae_automl = mae_automl/y_test_cut["akt. Leistung(kW)"].values.mean()

    # LOAD PARAMS.JSON
    # Load the JSON file
    save_dir_name = "AutoML_1"
    json_file_name = "params.json"
    json_file_path = os.path.join(save_dir_name, json_file_name)
    params_json = json.load(open(json_file_path))

    best_model = params_json["best_model"]

    # LOAD LEADEERBOARD.CSV
    leaderboard = pd.read_csv(os.path.join(save_dir_name, "leaderboard.csv"))
    # Find lowest value of "metric_value" column
    mae_cv = leaderboard["metric_value"].min()
    nmae_cv = mae_cv/y_train["akt. Leistung(kW)"].values.mean()

    # File paths
    wd_path = os.getcwd()
    detail_path = os.path.join(wd_path, "Best model predictions detail.png")
    overview_path = os.path.join(wd_path, "Best model predictions overview.png")
    autocorr_path = os.path.join(wd_path, "Autocorrelation.png")

    # Save a timestamp
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    # Create a textfile called "last_changed: timestamp"
    filename = "last_changed " + timestamp + ".txt"
    # Save file
    with open(filename, "w") as file:
        file.write("")

    return rmse_automl, mae_automl, r2_automl, nmae_automl, rmse_bl, mae_bl, r2_bl, nmae_bl, best_model, mae_cv, nmae_cv, detail_path, overview_path, autocorr_path

# Decision flow for choosing gapfilling mode
def decision_flow(hist_data, autocorrelation_peak, model_performance):
    """Decision flow for choosing the gap-filling mode.
    
    :param hist_data: Historical data availability.
    :type hist_data: bool
    :param autocorrelation_peak: Autocorrelation peak detection.
    :type autocorrelation_peak: bool
    :param model_performance: Model performance.
    :type model_performance: bool
    :return: Gap-filling mode.
    :rtype: str
    """

    if hist_data:
        if autocorrelation_peak:
            if model_performance:
                return "KI"  # AI-supported gap filling
            else:
                return "Vorwoche"  # Prior week
        else:
            return "Vorwoche" # Prior week
    else:
        return "SLP"  # Standard load profiles



# BOOKMARK: MAIN ##############################################################
###############################################################################
def main_routine():
    """Main routine for the machine learning (ML) analysis."""
    # Start Timer  
    start_time = time.time()

    # Save current working directory
    cwd_init = os.getcwd()

    # Clear the ml-models directory if overwrite enabled
    ml_models_dir = var_ml_models.ml_model_dir
    # Clear directory if overwrite is enabled
    if os.path.exists(ml_models_dir) and var_ml_models.overwrite_ml_models:
        shutil.rmtree(ml_models_dir)
    # Create directory if it doesn't exist
    if not os.path.exists(ml_models_dir):
        os.makedirs(ml_models_dir)

    # If a previous results file exists, read it in
    if os.path.exists(var_ml_models.results_xlsx_path):
        results_last = pd.read_excel(var_ml_models.results_xlsx_path, dtype={"Abnehmer": str, "Autokorrelation": str, "Detail": str, "Übersicht": str})

    # Read topology file
    topology = pd.read_excel(var_cons_list_analysis.topology_file, sheet_name="Knoten", dtype={"Abnehmer": str})
    # Filter for all active rows
    topology_filtered = topology[topology["aktiv"] == "ja"]
    # Filter for all rows where "Abnehmer" is not NaN
    topology_filtered = topology_filtered[topology_filtered["Abnehmer"].notna()]
    # Only get consumer IDs where "Hist. Daten existieren" is True (can't train models without historical data)
    cons_ids_hist_data = topology_filtered[topology_filtered["Hist. Daten existieren"] == True]["Abnehmer"].values
    # Get "Override" column for all consumers where "Hist. Daten existieren" is True
    override = topology_filtered[topology_filtered["Hist. Daten existieren"] == True]["Override"].values

    # Check if "Override" is set to ML for any consumer where no historical data exists
    topology_no_hist_data = topology[topology["Hist. Daten existieren"] == False]
    override_no_hist_data = topology_no_hist_data[topology_no_hist_data["Override"] == "ML"]
    if not override_no_hist_data.empty:
        raise ValueError("ML-modelle können nicht für Abnehmer ohne historische Daten erstellt werden." +
                            f"Bitte überprüfen Sie das Topologie-Excelsheet für Abnehmer {override_no_hist_data['Abnehmer'].values}")

    # Create lists to store results
    cons_id_list, rmse_automl_list, mae_automl_list, r2_automl_list, mape_automl_list, rmse_bl_list, \
        mae_bl_list, r2_bl_list, mape_bl_list, best_model_list, mae_cv_list, nmae_cv_list, \
        detail_path_list, overview_path_list, autocorr_path_list, \
        autocorr_peaks_exist_list = [[] for i in range(16)]

    dummy1, rmse_automl, mae_automl, r2_automl, mape_automl, rmse_bl, \
        mae_bl, r2_bl, mape_bl, mae_cv, nmae_cv, dummy2, autocorr_peaks_exist, \
        best_model, autocorr_path, detail_path, overview_path, dummy3 = [None for i in range(18)]
        
    # Iterate through list
    for ind, cons_id in enumerate(cons_ids_hist_data):
        print_green("Analysiere Abnehmer " + cons_id)

        # FOLDER STRUCTURE ####################################################
        folder_path = os.path.join(ml_models_dir, cons_id)
        # If overwrite is disabled, check if folder exists
        if not var_ml_models.overwrite_ml_models:
            # Check if folder exists
            if os.path.exists(folder_path):
                # Check if there's a folder named "AutoML_1"
                automl_folder = os.path.join(folder_path, "AutoML_1")
                if os.path.exists(automl_folder):
                    # Check if there's a text file called "last_changed"
                    last_changed = [f for f in os.listdir(folder_path) if "last_changed" in f]
                    if last_changed:
                        # Get the last changed date
                        last_changed = last_changed[0]
                        last_changed = last_changed.replace("last_changed ", "")
                        last_changed = last_changed.replace(".txt", "")
                        last_changed = datetime.datetime.strptime(last_changed, "%Y-%m-%d_%H-%M-%S")
                        last_changed = datetime.datetime.strftime(last_changed, "%d.%m.%Y um %H:%M:%S")
                        # Get results from previous results file
                        # Extract row from results_last
                        results_last_row = results_last[results_last["Abnehmer"] == cons_id]
                        dummy1, rmse_automl, mae_automl, r2_automl, mape_automl, rmse_bl, \
                        mae_bl, r2_bl, mape_bl, mae_cv, nmae_cv, dummy2, autocorr_peaks_exist, \
                        best_model, autocorr_path, detail_path, overview_path, dummy3 = [results_last_row[col].values[0] for col in results_last.columns]
                        # Stupid workaround bc pandas can't read hyperlinks from excel
                        autocorr_path = os.path.join(folder_path, "Autocorrelation.png")
                        detail_path = os.path.join(folder_path, "Best model predictions detail.png")
                        overview_path = os.path.join(folder_path, "Best model predictions overview.png")
                        # Append results to lists
                        cons_id_list.append(cons_id)
                        rmse_automl_list.append(rmse_automl)
                        mae_automl_list.append(mae_automl)
                        r2_automl_list.append(r2_automl)
                        mape_automl_list.append(mape_automl)
                        rmse_bl_list.append(rmse_bl)
                        mae_bl_list.append(mae_bl)
                        r2_bl_list.append(r2_bl)
                        mape_bl_list.append(mape_bl)
                        best_model_list.append(best_model)
                        mae_cv_list.append(mae_cv)
                        nmae_cv_list.append(nmae_cv)
                        detail_path_list.append(detail_path)
                        overview_path_list.append(overview_path)
                        autocorr_path_list.append(autocorr_path)
                        autocorr_peaks_exist_list.append(autocorr_peaks_exist)  
                        # Skip this consumer
                        print(f"Für diesen Abnehmer existiert bereits ein Modell vom {last_changed}. Überspringe.")
                        continue
                    else:
                        # Reset folder
                        shutil.rmtree(folder_path)
                        os.makedirs(folder_path)
                else:
                    # Reset folder
                    shutil.rmtree(folder_path)
                    os.makedirs(folder_path)
            else:
                # Create folder
                os.makedirs(folder_path)
        # If overwrite is enabled, create folder
        else:
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path)
                os.makedirs(folder_path)
            else:
                os.makedirs(folder_path)


        # Set this folder as the working directory
        os.chdir(folder_path)

        # DO ML STUFF #########################################################
        # Check autocorrelation functions
        autocorr_peaks_exist = preliminary_analysis(cons_id)

        if autocorr_peaks_exist or var_ml_models.force_ml_models or override[ind] == "ML":
            rmse_automl, mae_automl, r2_automl, mape_automl, rmse_bl, mae_bl, \
                r2_bl, mape_bl, best_model, mae_cv, nmae_cv, detail_path, overview_path, \
                autocorr_path= create_ml_models(cons_id)
        else:
            print(f"Für Abnehmer {cons_id} wird kein ML-Modell erstellt.")
            rmse_automl, mae_automl, r2_automl, mape_automl, rmse_bl, mae_bl, \
                r2_bl, mape_bl, best_model, mae_cv, nmae_cv, detail_path, \
                overview_path = [None for i in range(13)]
            autocorr_path = os.path.join(os.getcwd(), "Autocorrelation.png")

            # try:

            # except:
            #     print_red(f"Abnehmer {cons_id}: Fehler beim Erstellen der ML-Modelle.")
            #     rmse_automl, mae_automl, r2_automl, mape_automl, rmse_bl, mae_bl, \
            #         r2_bl, mape_bl, best_model, mae_cv, nmae_cv, detail_path, overview_path, \
            #         autocorr_path, echelle_path = [None for i in range(15)]
            #     continue
        
        # Append results to lists
        cons_id_list.append(cons_id)
        rmse_automl_list.append(rmse_automl)
        mae_automl_list.append(mae_automl)
        r2_automl_list.append(r2_automl)
        mape_automl_list.append(mape_automl)
        rmse_bl_list.append(rmse_bl)
        mae_bl_list.append(mae_bl)
        r2_bl_list.append(r2_bl)
        mape_bl_list.append(mape_bl)
        best_model_list.append(best_model)
        mae_cv_list.append(mae_cv)
        nmae_cv_list.append(nmae_cv)
        detail_path_list.append(detail_path)
        overview_path_list.append(overview_path)
        autocorr_path_list.append(autocorr_path)
        autocorr_peaks_exist_list.append(autocorr_peaks_exist)        

        # Set working directory back to initial
        os.chdir(cwd_init)

        # Compute difference between NMAE of AutoML and baseline
        nmae_diff = [mape_automl_list[i] - mape_bl_list[i] \
                        if mape_automl_list[i] is not None and mape_bl_list[i] is not None else None \
                        for i in range(len(mape_automl_list))]

        # Save results to Excel
        results = pd.DataFrame({"Abnehmer": cons_id_list,
                                "RMSE AutoML (Validierungsset)": rmse_automl_list,
                                "MAE AutoML (Validierungsset)": mae_automl_list,
                                "R2 AutoML (Validierungsset)": r2_automl_list,
                                "NMAE AutoML (Validierungsset)": mape_automl_list,
                                "RMSE Baseline (Validierungsset)": rmse_bl_list,
                                "MAE Baseline (Validierungsset)": mae_bl_list,
                                "R2 Baseline (Validierungsset)": r2_bl_list,
                                "NMAE Baseline (Validierungsset)": mape_bl_list,
                                "MAE Kreuzvalidierung": mae_cv_list,
                                "NMAE Kreuzvalidierung": nmae_cv_list,
                                "NMAE Differenz AutoML-Baseline (Validierungsset)": nmae_diff,
                                "Peaks bei 24 und 48 h": autocorr_peaks_exist_list,
                                "Bestes modell": best_model_list,
                                "Autokorrelation": autocorr_path_list,
                                "Detail": detail_path_list,
                                "Übersicht": overview_path_list})



        # Determine gapfilling mode to use for each consumer
        # Iterate through topology file
        for row in topology_filtered.iterrows():
            # Get consumer ID
            cons_id = row[1]["Abnehmer"]
            # If cons_id is NaN, skip
            if pd.isna(cons_id):
                continue

            # Check if consumer has historical data
            hist_data_good = row[1]["Hist. Daten existieren"]
            # Convert to bool
            hist_data_good = hist_data_good == 1.0
            # Check if consumer has autocorrelation peaks
            results_current_row = results[results["Abnehmer"] == cons_id]
            try:
                autocorr_good = results_current_row["Peaks bei 24 und 48 h"].values[0]
            except:
                autocorr_good = False
            # Check if consumer has good model performance
            try:
                nmae_diff_cons = results_current_row["NMAE Differenz AutoML-Baseline (Validierungsset)"].values[0]
            except:
                nmae_diff_cons = None
            if pd.isna(nmae_diff_cons):
                nmae_good = False
            else:
                nmae_good = nmae_diff_cons < -0.05
            
            gapfilling_mode = decision_flow(hist_data_good,autocorr_good,nmae_good)

            # Write gapfilling mode to topology df
            topology.loc[topology["Abnehmer"] == cons_id, "Lückenfüllung"] = gapfilling_mode

        # WRITE TOPOLOGY FILE #########################################################
        ###############################################################################
        # Save gapfilling information to topology file
        topo_wb = openpyxl.load_workbook(var_load_profiles.topology_file)
        topo_ws = topo_wb["Knoten"]
        # Write "Lückenfüllung" to column X of the topology file
        for index, row in topology.iterrows():
            topo_ws[f"X{index+2}"] = row["Lückenfüllung"]

        topo_wb.save(var_load_profiles.topology_file)

        # Merge "Lückenfüllung" into results on "Abnehmer"
        results = pd.merge(results, topology[["Abnehmer", "Lückenfüllung"]], on="Abnehmer", how="left")

        # Save results to Excel
        results.to_excel(var_ml_models.results_xlsx_path, index=False)
        format_workbook(var_ml_models.results_xlsx_path)

    # End Timer
    end_time = time.time()
    # Calculate time difference
    time_diff = end_time - start_time
    # Print time difference in hours
    print(f"Time elapsed: {time_diff/3600} hours")

    Beep(440, 1000)

if __name__ == "__main__":
    main_routine()



